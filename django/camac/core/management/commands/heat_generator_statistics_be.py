import argparse
import itertools
from datetime import datetime

import openpyxl
from caluma.caluma_form.models import DynamicOption, Option
from caluma.caluma_workflow.models import Case
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from tqdm import tqdm

from camac.caluma.api import CalumaApi

caluma_api = CalumaApi()

DOSSIER_NUMBER = "Dossier Nummer"
EBAU_NUMBER = "eBau Nummer"
SUBMIT_DATE = "Eingabedatum"
STREET = "Strasse/Flurname"
NR = "Nr."
PLZ = "PLZ"
LOCATION = "Ort"
MUNICIPALITY = "Gemeinde"
GWR_EGID = "GWR-EGID"
DECISION = "Entscheid"
BUILDING_CATEGORY = "Gebäudekategorie"
CREATION_YEAR = "Erstellungsjahr"
EXISING_ENERGY_SOURCE = "Bestehende Wärmeerzeuger - Energieträger/System"
EXISTING_HEAT_CAPACITY = "Bestehende Wärmeerzeuger - Heizleistung in kW"
EXISTING_WATER_WARMING = "Bestehende Wärmeerzeuger - Warmwassererwärmung"
EXISTING_SOLAR_ENERGY = "Bestehende Wärmeerzeuger - Solare Energienutzung"
NEW_ENERGY_SOURCE = "Neue Wärmeerzeuger - Energieträger/System"
NEW_HEAT_CAPACITY = "Neue Wärmeerzeuger - Heizleistung in kW"
NEW_WATER_WARMING = "Neue Wärmeerzeuger - Warmwassererwärmung"
EXISTING_REQUIREMENTS_ENERGY_SOURCE = (
    "Neue Wärmeerzeuger mit Anforderungen - Energieträger/System"
)
EXISTING_REQUIREMENTS_HEAT_CAPACITY = (
    "Neue Wärmeerzeuger mit Anforderungen - Heizleistung in kW"
)
EXISTING_REQUIREMENTS_WATER_WARMING = (
    "Neue Wärmeerzeuger mit Anforderungen - Warmwassererwärmung"
)
REPLACEMENT_OF = "Ersatz von"


def valid_date(date_string):
    """Validate and convert a string in 'dd.mm.YYYY' format to a date object."""

    try:
        return datetime.strptime(date_string, "%d.%m.%Y").date()
    except ValueError:
        raise argparse.ArgumentTypeError(
            f"Not a valid date: '{date_string}'. Use the 'dd.mm.YYYY' format."
        )


class Command(BaseCommand):
    help = """Create a xlsx file with heat generator statistics about instances."""

    def add_arguments(self, parser):
        parser.add_argument(
            "--date-from",
            type=valid_date,
            help="Start date in dd.mm.YYYY format (e.g. 05.06.2026)",
        )
        parser.add_argument(
            "--date-to",
            type=valid_date,
            help="End date in dd.mm.YYYY format (e.g. 10.06.2026)",
        )

    def _get_val_with_v2_fallback(self, row, key):
        if not row:
            return None
        return row.get(key) or row.get(f"{key}-v2")

    def _get_option_label(self, slug):
        if not slug:
            return "-"
        try:
            return Option.objects.get(slug=slug).label.de or "-"
        except Option.DoesNotExist:
            return "-"

    def _add_multiple_choice_values(self, case, slugs, entry, fieldname):
        data = []
        answers = [a for a in case.document.answers.all() if a.question_id in slugs]
        for answer in answers:
            for value in answer.value:
                options = answer.question.options.filter(slug=value)
                if options:
                    data.append(options[0].label.de)
        entry[fieldname] = ", ".join(data)

    def _fetch_cases(self, options):
        if options.get("date_from"):
            date_from_str = f"{options['date_from'].strftime('%Y-%m-%d')}T00:00:00+0000"
        if options.get("date_to"):
            date_to_str = f"{options['date_to'].strftime('%Y-%m-%d')}T23:59:59+0000"
        self.stdout.write(
            "Fetching cases from the database... (this may take a minute)"
        )
        cases_queryset = (
            Case.objects.filter(
                document__form__slug__in=[
                    "heat-generator",
                    "heat-generator-v2",
                    "heat-generator-v3",
                ],
                **{
                    "meta__submit-date__range": (
                        date_from_str,
                        date_to_str,
                    )
                },
            )
            .select_related("instance", "document")
            .prefetch_related(
                "work_items__document__answers__question__options",
                "document__answers__question__options",
            )
            .order_by("instance__pk")
        )

        self.stdout.write(
            f"Successfully fetched {cases_queryset.count()} cases. Starting processing..."
        )
        return cases_queryset

    def _get_decision(self, case):
        decision_item = case.work_items.filter(
            task="decision", status="completed"
        ).first()
        if not decision_item:
            return "-"

        answer = decision_item.document.answers.filter(
            question_id="decision-decision-assessment"
        ).first()

        if not answer:
            return "-"

        option = answer.question.options.filter(slug=answer.value).first()
        return option.label.de if option and option.label.de else "-"

    def _build_base_entry(self, case, flat_answers):
        entry = {
            DOSSIER_NUMBER: case.instance.pk,
            EBAU_NUMBER: case.meta.get("ebau-number", "-"),
            SUBMIT_DATE: datetime.strptime(
                case.meta.get("submit-date"), "%Y-%m-%dT%H:%M:%S%z"
            ).strftime("%d.%m.%Y"),
            DECISION: self._get_decision(case),
        }

        self._add_multiple_choice_values(
            case, ["heat-generator-category"], entry, "Gebäudekategorie"
        )
        self._add_multiple_choice_values(
            case,
            [
                "heat-generator-substituted-type",
                "heat-generator-substituted-type-v2",
            ],
            entry,
            "Ersatz von",
        )

        creation_year_answer = case.document.answers.filter(
            question_id="heat-generator-year"
        ).first()
        entry[CREATION_YEAR] = (
            creation_year_answer.value if creation_year_answer else "-"
        )

        entry[STREET] = flat_answers.get("strasse-flurname", "-")
        entry[NR] = flat_answers.get("nr", "-")
        entry[PLZ] = next(
            (
                val
                for key, val in flat_answers.items()
                if key.startswith("plz-grundstueck")
            ),
            "-",
        )
        entry[LOCATION] = flat_answers.get("ort-grundstueck", "-")
        entry[GWR_EGID] = flat_answers.get("gwr-egid", "-")

        gemeinde_slug = flat_answers.get("gemeinde")
        if gemeinde_slug:
            municipality_opt = DynamicOption.objects.filter(
                slug=gemeinde_slug, document_id=case.document.pk
            ).first()
            entry[MUNICIPALITY] = municipality_opt.label.de if municipality_opt else "-"
        else:
            entry[MUNICIPALITY] = "-"

        return entry

    def _process_tables(self, flat_answers, base_entry, data):
        existing_table = (flat_answers.get("heat-generator-existing") or []) + (
            flat_answers.get("heat-generator-existing-v2") or []
        )
        new_table = (flat_answers.get("heat-generator-new") or []) + (
            flat_answers.get("heat-generator-new-v2") or []
        )
        req_table = (flat_answers.get("heat-generator-new-with-requirements") or []) + (
            flat_answers.get("heat-generator-new-with-requirements-v2") or []
        )

        if not existing_table and not new_table and not req_table:
            data.append(base_entry)
            return

        safe_existing = existing_table or [{}]
        safe_new = new_table or [{}]
        safe_req = req_table or [{}]

        for existing_row, new_row, req_row in itertools.product(
            safe_existing, safe_new, safe_req
        ):
            row_entry = base_entry.copy()

            if existing_row:
                self._apply_existing_row(existing_row, row_entry)
            if new_row:
                self._apply_new_row(new_row, row_entry)
            if req_row:
                self._apply_req_row(req_row, row_entry)

            data.append(row_entry)

    def _apply_existing_row(self, row, row_entry):
        raw_source = self._get_val_with_v2_fallback(
            row, "heat-generator-energy-source-existing"
        )
        raw_cap = self._get_val_with_v2_fallback(row, "heat-generator-capacity")
        raw_water = self._get_val_with_v2_fallback(row, "heat-generator-water-heating")
        raw_solar = self._get_val_with_v2_fallback(
            row, "heat-generator-solar-energy-usage"
        )

        row_entry[EXISING_ENERGY_SOURCE] = self._get_option_label(raw_source)
        row_entry[EXISTING_HEAT_CAPACITY] = raw_cap or "-"
        row_entry[EXISTING_WATER_WARMING] = self._get_option_label(raw_water)
        row_entry[EXISTING_SOLAR_ENERGY] = self._get_option_label(raw_solar)

    def _apply_new_row(self, row, row_entry):
        raw_source_new = self._get_val_with_v2_fallback(
            row, "heat-generator-energy-source-new"
        )
        raw_cap_new = self._get_val_with_v2_fallback(row, "heat-generator-capacity")
        raw_water_new = self._get_val_with_v2_fallback(
            row, "heat-generator-water-heating"
        )

        row_entry[NEW_ENERGY_SOURCE] = self._get_option_label(raw_source_new)
        row_entry[NEW_HEAT_CAPACITY] = raw_cap_new or "-"
        row_entry[NEW_WATER_WARMING] = self._get_option_label(raw_water_new)

    def _apply_req_row(self, row, row_entry):
        raw_source_req = self._get_val_with_v2_fallback(
            row, "heat-generator-energy-source-new-with-requirements"
        )
        raw_cap_req = self._get_val_with_v2_fallback(row, "heat-generator-capacity")
        raw_water_req = self._get_val_with_v2_fallback(
            row, "heat-generator-water-heating"
        )

        row_entry[EXISTING_REQUIREMENTS_ENERGY_SOURCE] = self._get_option_label(
            raw_source_req
        )
        row_entry[EXISTING_REQUIREMENTS_HEAT_CAPACITY] = raw_cap_req or "-"
        row_entry[EXISTING_REQUIREMENTS_WATER_WARMING] = self._get_option_label(
            raw_water_req
        )

    @transaction.atomic
    def handle(self, *args, **options):

        cases_queryset = self._fetch_cases(options)

        data = []
        for case in tqdm(cases_queryset, desc="Processing cases", unit="case"):
            flat_answers = case.document.flat_answer_map()
            base_entry = self._build_base_entry(case, flat_answers)
            self._process_tables(flat_answers, base_entry, data)

        self.generate_excel(data)
        self.stdout.write(
            self.style.SUCCESS(f"Successfully exported {cases_queryset.count()}.")
        )

    def generate_excel(self, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WEU Statistics"

        fieldnames = [
            DOSSIER_NUMBER,
            EBAU_NUMBER,
            SUBMIT_DATE,
            STREET,
            NR,
            PLZ,
            LOCATION,
            MUNICIPALITY,
            GWR_EGID,
            DECISION,
            BUILDING_CATEGORY,
            CREATION_YEAR,
            EXISING_ENERGY_SOURCE,
            EXISTING_HEAT_CAPACITY,
            EXISTING_WATER_WARMING,
            EXISTING_SOLAR_ENERGY,
            NEW_ENERGY_SOURCE,
            NEW_HEAT_CAPACITY,
            NEW_WATER_WARMING,
            EXISTING_REQUIREMENTS_ENERGY_SOURCE,
            EXISTING_REQUIREMENTS_HEAT_CAPACITY,
            EXISTING_REQUIREMENTS_WATER_WARMING,
            REPLACEMENT_OF,
        ]

        bold_font = Font(bold=True)
        right_alignment = Alignment(horizontal="right")

        # Write and style the header row
        ws.append(fieldnames)

        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = right_alignment

        # Write and style the data rows
        for row_index, row_dict in enumerate(data, start=2):
            row_values = []
            for field in fieldnames:
                val = row_dict.get(field)
                if val in (None, "", []):
                    row_values.append("-")
                else:
                    row_values.append(val)

            ws.append(row_values)

            for cell in ws[row_index]:
                cell.alignment = right_alignment

        self._adjust_column_widths(ws)

        # Add the AutoFilter
        ws.auto_filter.ref = ws.dimensions

        # Highlight duplicate Dossier Numbers in Column A
        last_row = len(data) + 1
        if last_row >= 2:
            red_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            red_text = Font(color="9C0006")

            dxf = DifferentialStyle(font=red_text, fill=red_fill)

            rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=True)

            ws.conditional_formatting.add(f"A2:A{last_row}", rule)

        # Freeze the header row so it's always visible
        ws.freeze_panes = "A2"

        wb.save("weu_statistic.xlsx")

    def _adjust_column_widths(self, ws):
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter

            for cell in col:
                try:
                    cell_text = str(cell.value) if cell.value is not None else ""

                    if cell.row == 1:
                        current_length = len(cell_text) + 6
                    else:
                        current_length = len(cell_text)

                    if current_length > max_length:
                        max_length = current_length
                except Exception:
                    pass

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
