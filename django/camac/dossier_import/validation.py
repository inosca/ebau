import zipfile
from enum import Enum
from typing import List

import openpyxl
from django.conf import settings
from django.utils import timezone
from django.utils.module_loading import import_string
from django.utils.translation import gettext as _
from rest_framework.exceptions import ValidationError

from camac.dossier_import import messages
from camac.dossier_import.loaders import InvalidImportDataError
from camac.dossier_import.models import DossierImport

from .config.common import mimetypes
from .loaders import XlsxFileDossierLoader
from .messages import MessageCodes


class TargetStatus(Enum):
    SUBMITTED = "SUBMITTED"
    APPROVED = "APPROVED"
    REJECTED = "REJECTED"
    WRITTEN_OFF = "WRITTEN OFF"
    DONE = "DONE"


REQUIRED_COLUMNS = ["proposal", "status", "submit-date"]


def verify_source_file(source_file: str) -> str:
    """
    Verify source file is valid archive and has content.

    Verfication here goes pretty far. But since it's about rather big archive files
    we do not want to store any file that is not going to survive the first
    step of processing anyway.

    Failing to validate source_file results in an APIError i. e. bad request

    The following steps are required:
     - archive must be a ZIP file
     - the archive must contain a dossiers.xlsx
     - the dossiers.xlsx must in fact be a XLSX file
    """
    if source_file is None:
        raise ValidationError(_("To start an import please upload a file."))

    try:
        archive = zipfile.ZipFile(source_file)
    except zipfile.BadZipfile:
        raise ValidationError(_("Uploaded file is not a valid .zip file"))
    try:
        metadata = archive.open("dossiers.xlsx")
    except KeyError:
        raise ValidationError(
            _("No metadata file 'dossiers.xlsx' found in uploaded archive.")
        )
    try:
        openpyxl.load_workbook(metadata, data_only=True)
    except zipfile.BadZipfile:
        raise ValidationError(
            _("Metadata file `dossiers.xlsx` is not a valid .xlsx file.")
        )
    return source_file


def validate_attachments(archive: zipfile.ZipFile, dossier_ids: List[str]):
    dirs = {
        filename.split("/")[0]
        for filename in archive.namelist()
        if "/" in filename and len(filename.split("/")) < 3
    }
    orphan_dirs = sorted(list(dirs - set(dossier_ids)))
    result = []
    if orphan_dirs:
        result.append(
            _(
                "%(count)i document folders were not found in the metadata file and will not be imported:\n%(entries)s"
            )
            % dict(count=len(orphan_dirs), entries=", ".join(orphan_dirs))
        )
    dossiers_without_documents = list(set(dossier_ids) - dirs)
    if dossiers_without_documents:
        result.append(
            _("%(count)i dossiers have no document folder.")
            % dict(count=len(dossiers_without_documents))
        )
    return result


def get_attachment_validation_stats(archive: zipfile.ZipFile, dossier_ids: List[str]):
    return sum(
        [
            1
            for filename in archive.namelist()
            if filename.split("/")[0] in dossier_ids
            and not filename.endswith("/")
            and mimetypes.guess_type(filename)
        ]
    )


def find_cell_by_value(value, cells):
    return next((cell for cell in cells if cell.value == value), None)


def get_dossier_cell_by_column(dossier_id, col_name, worksheet):
    cell = find_cell_by_value(
        col_name.upper(),
        worksheet[1],
    )
    if not cell:  # pragma: no cover
        return
    for row in worksheet.iter_rows(min_row=2):
        match = find_cell_by_value(dossier_id, row)
        if match:
            # cells return their original row from the sheet.
            # to get the accurate index the headings row has to be
            # subtracted as well
            return row[cell.col_idx - 1]


def _get_first_row(worksheet):
    return [heading for heading in list(worksheet.rows)[0]]


def _validate_date_fields(
    dossier_id, dossier_messages, worksheet, allow_delete=False
) -> bool:
    # openpyxl may provide data from cells based on the cell formatting.
    # Specs define that we accept values in the XlsxFileDossierLoader.date_format.
    # something similar to "03.04.2001".
    # Uplodaded documents can feature those columns formatted as datetime as well as
    # strings with correctly entered data. In that case we should try to parse that
    # value or return instructions.
    valid = True
    for date_column in [
        heading
        for heading in _get_first_row(worksheet)
        if heading.value and heading.value.endswith("-DATE")
    ]:
        if (
            date := get_dossier_cell_by_column(dossier_id, date_column.value, worksheet)
        ) and date.value:
            # We'll use these for handling messages below repeatedly.
            # Individual messages require `code`, `detail` and `level`
            # set separately.
            message_defaults = {
                "dossier_id": dossier_id,
                "field_name": date_column.value.lower(),
                "messages": dossier_messages,
            }

            if date.value == settings.DOSSIER_IMPORT["DELETE_KEYWORD"]:
                if allow_delete:
                    messages.append_or_update_dossier_message(
                        code=MessageCodes.VALUE_DELETED,
                        detail="Date will be deleted",
                        level=messages.Severity.INFO.value,
                        **message_defaults,
                    )

                else:
                    messages.append_or_update_dossier_message(
                        code=MessageCodes.FIELD_VALIDATION_ERROR,
                        detail=(
                            f"Ignoring {date.value} for {date_column.value.lower()} because ",
                            "Deletion is not supported for new dossiers",
                        ),
                        level=messages.Severity.WARNING.value,
                        **message_defaults,
                    )

                continue

            if not isinstance(date.value, timezone.datetime):
                try:
                    date = timezone.datetime.strptime(
                        date.value, XlsxFileDossierLoader.date_format
                    )
                except ValueError:
                    messages.append_or_update_dossier_message(
                        detail=str(date.value),
                        code=MessageCodes.DATE_FIELD_VALIDATION_ERROR.value,
                        level=messages.Severity.ERROR.value,
                        **message_defaults,
                    )
                    valid = False
    return valid


def _validate_existing_dossier(dossier_id, dossier_msgs, worksheet):
    messages.append_or_update_dossier_message(
        dossier_id=dossier_id,
        field_name="ID",
        detail=dossier_id,
        code=MessageCodes.UPDATE_DOSSIER.value,
        messages=dossier_msgs,
        level=messages.Severity.WARNING.value,
    )
    # reimporting does not require the "STATUS" column to be present, therefore
    # we do cannot ta
    if (
        status := get_dossier_cell_by_column(dossier_id, "status", worksheet)
    ) and status.value:
        messages.append_or_update_dossier_message(
            dossier_id=dossier_id,
            field_name="STATUS",
            detail=_("STATUS will be ignored when updating the dossier"),
            code=MessageCodes.UPDATE_DOSSIER.value,
            messages=dossier_msgs,
            level=messages.Severity.INFO.value,
        )

    _validate_date_fields(dossier_id, dossier_msgs, worksheet, allow_delete=True)
    return True


def _validate_new_dossier(dossier_id, dossier_msgs, worksheet):
    # New dossiers require additional columns to the `ID` column.
    # If we encounter this we rather immediately raise to avoid complex
    # communication.
    _raise_for_missing_columns(
        worksheet, *[col_name.upper() for col_name in REQUIRED_COLUMNS]
    )
    valid = True
    # check that status is not empty and also valid
    if (
        (status := get_dossier_cell_by_column(dossier_id, "status", worksheet))
        and status.value
        and status.value not in [e.value for e in TargetStatus]
    ):
        messages.append_or_update_dossier_message(
            dossier_id,
            "status",
            status.value,
            MessageCodes.STATUS_CHOICE_VALIDATION_ERROR.value,
            dossier_msgs,
            level=messages.Severity.ERROR.value,
        )
        valid = False
    # check submit date exists in newly imported dossiers
    if (
        cell := get_dossier_cell_by_column(dossier_id, "submit-date".upper(), worksheet)
    ) and not cell.value:
        messages.append_or_update_dossier_message(
            dossier_id,
            "submit-date",
            None,
            MessageCodes.MISSING_REQUIRED_VALUE_ERROR.value,
            dossier_msgs,
            level=messages.Severity.ERROR.value,
        )
        valid = False
    valid = _validate_date_fields(
        dossier_id, dossier_msgs, worksheet, allow_delete=False
    )
    return valid


def _raise_for_missing_columns(worksheet, *columns):
    if any(
        (
            missing_columns := [
                col
                for col in columns
                if find_cell_by_value(col, _get_first_row(worksheet)) is None
            ]
        )
    ):
        raise InvalidImportDataError(
            _(
                "Meta data file in archive is missing required columns %(missing)s. Required columns %(required)s"
            )
            % dict(
                required=", ".join(columns),
                missing=", ".join([col for col in missing_columns]),
            )
        )


def validate_zip_archive_structure(instance_pk, clean_on_fail=True) -> DossierImport:
    """
    ZIP archive validation.

    scans the archive and best guesses the outcome of actually importing it.
    """
    dossier_import = DossierImport.objects.get(pk=instance_pk)

    configured_writer_cls = import_string(settings.DOSSIER_IMPORT["WRITER_CLASS"])
    writer = configured_writer_cls(
        user_id=dossier_import.user.pk,
        group_id=dossier_import.group.pk,
        location_id=dossier_import.location and dossier_import.location.pk,
    )

    archive = dossier_import.get_archive()
    data_file = archive.open("dossiers.xlsx")
    try:
        work_book = openpyxl.load_workbook(data_file, data_only=True)
    except zipfile.BadZipfile:
        raise InvalidImportDataError(
            _("Meta data file in archive is corrupt or not a valid .xlsx file.")
        )
    worksheet = work_book.worksheets[0]
    if extra := set([cell.value for cell in _get_first_row(worksheet)]) - set(
        [e.value for e in XlsxFileDossierLoader.Column]
    ):
        sorted_missing_columns = [str(x) for x in (extra - set(["None", None]))]
        dossier_import.messages["validation"]["summary"]["warning"].append(
            _(
                "Found unknown columns: %(extra)s. Those columns will be ignored while importing."
            )
            % dict(extra="; ".join(sorted_missing_columns))
        )

    # collect all messages for every dossier in a list
    dossier_msgs = []

    _raise_for_missing_columns(worksheet, "ID")

    # get the list of IDs in the file
    id_column = find_cell_by_value("ID", _get_first_row(worksheet)).col_idx - 1
    dossier_ids = [
        row[id_column].value
        for row in list(worksheet.rows)[1:]
        if row[id_column].value is not None
    ]

    # exclude and report duplicate IDs
    dupes = set([d for d in dossier_ids if dossier_ids.count(d) > 1])
    for dupe in dupes:
        messages.append_or_update_dossier_message(
            dossier_id=dupe,
            field_name="id",
            detail=f"Multiple rows found with id {dupe}. Ignoring",
            code=MessageCodes.DUPLICATE_IDENTFIER_ERROR.value,
            messages=dossier_msgs,
        )
        dossier_ids.remove(dupe)

    # Each dossier ID may be discarded for various reasons and based on
    # different criteria depending on the import status as a new or reimported
    # dossier.
    valid_dossier_ids = []
    for dossier_id in set(dossier_ids):
        valid = True

        # NEW dossier: perform validations for a new dossier
        if not writer.existing_dossier(dossier_id):
            valid = _validate_new_dossier(dossier_id, dossier_msgs, worksheet)
            if valid:
                valid_dossier_ids.append(dossier_id)
            continue

        # otherwise handle reimport
        valid = _validate_existing_dossier(
            dossier_id,
            dossier_msgs,
            worksheet,
        )
        if valid:
            valid_dossier_ids.append(dossier_id)

    # handle messages collected while validating the rows for the details
    # section.
    for msg in dossier_msgs:
        messages.update_messages_section_detail(
            message=msg, section="validation", dossier_import=dossier_import
        )

    dossier_import = messages.update_summary(dossier_import)
    dossier_import.messages["validation"]["summary"]["warning"] += validate_attachments(
        archive, valid_dossier_ids
    )
    dossier_import.messages["validation"]["summary"]["stats"] = {
        "attachments": get_attachment_validation_stats(archive, dossier_ids),
        "dossiers": len(valid_dossier_ids),
    }

    dossier_import.messages["validation"]["completed"] = timezone.localtime().strftime(
        "%Y-%m-%dT%H:%M:%S%z"
    )
    dossier_import.save()

    dossier_import.status = dossier_import.IMPORT_STATUS_VALIDATION_SUCCESSFUL
    if dossier_import.messages["validation"]["summary"]["error"]:
        dossier_import.status = dossier_import.IMPORT_STATUS_VALIDATION_FAILED
        if clean_on_fail:
            dossier_import.source_file.delete()
    dossier_import.save()

    return dossier_import
