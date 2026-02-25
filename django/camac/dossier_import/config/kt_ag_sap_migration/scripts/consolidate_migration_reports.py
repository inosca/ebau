#!/usr/bin/env python3
"""
Migration Report Consolidation Script.

This script consolidates CSV files from migration report directories into Excel files.
Each Excel file contains worksheets for different report types (document_export,
document_import, dossier_import) and combines data from multiple CSV files.

Key features:
- Processes input directories in chronological order (oldest first)
- Uses centralized nested dict structure: Segment → Gemeinde → Gesuch-ID → [Dateiname] → List[rows]
- Applies all filters to the oldest import directory, only DOSSIER_IMPORT_FILTERS to others
- Generates all.xlsx with all rows from all migrations
- Generates all-consolidated.xlsx with first error-free row per key or last row if none exist
- Handles specificity filtering: excludes generic error rows when specific rows exist
- Sorts output by Segment, Gemeinde, Gesuch-ID, Dateiname, Import-Zeit
- Preserves multi-line cells in Excel output
- Automatically adjusts column widths based on content
- Adds autofilter to header rows for easy data filtering

Usage:
    python consolidate_migration_reports.py INPUT_PATH1 [INPUT_PATH2 ...] -o OUTPUT_DIR

Example:
    python consolidate_migration_reports.py /path/to/migration1 /path/to/migration2 -o /path/to/output
"""

import argparse
import csv
import datetime
import logging
import re
import sys
import traceback
from pathlib import Path
from typing import Dict, List, Optional, Pattern, Set, Tuple

import openpyxl

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(Path("consolidate_reports.log").resolve()),
    ],
)
logger = logging.getLogger(__name__)

# Global regex filters for each report type
# These can be modified to filter out rows based on specific patterns
#
# How to use regex filters:
# 1. Each filter is a compiled regular expression pattern
# 2. For each row in a CSV file, all cells are joined into a single line (with newlines replaced by spaces)
# 3. If any filter pattern matches this line, the entire row is excluded from the output
# 4. In the Excel output, multi-line cells are preserved as they were in the original CSV
#
# There are two sets of filters for each report type:
# - FIRST_TIME_*_FILTERS: Applied only to the oldest chronological migration source directory
# - ALWAYS_*_FILTERS: Applied to all migration source directories
#
# Example filters:
# FIRST_TIME_DOCUMENT_EXPORT_FILTERS = [
#     re.compile(r"error", re.IGNORECASE),  # Filter out rows containing "error" (case insensitive)
#     re.compile(r"^test", re.IGNORECASE),  # Filter out rows starting with "test"
# ]
#
# Note: The original multi-line structure of cells is preserved in the Excel output,
# only the regex matching process temporarily replaces newlines with spaces.

# Filters applied only to the oldest chronological migration source directory
FIRST_TIME_DOCUMENT_EXPORT_FILTERS: List[Pattern] = [
    # re.compile(r"finished raised ChunkedEncodingError"),
    # re.compile(r"replication ended with status"),
]
FIRST_TIME_DOCUMENT_IMPORT_FILTERS: List[Pattern] = [
    # re.compile(r"An unexpected error occurred while uploading the file"),
    # re.compile(r"connection timeout expired"),
    # re.compile(r"server closed the connection unexpectedly"),
    # re.compile(r"the connection is closed"),
    # re.compile(r"Keyword EBPA-.* not found in database."),
    # re.compile(r"in category intern.', code="),
]
FIRST_TIME_DOSSIER_IMPORT_FILTERS: List[Pattern] = []

# Filters applied to all migration source directories
ALWAYS_DOCUMENT_EXPORT_FILTERS: List[Pattern] = []
ALWAYS_DOCUMENT_IMPORT_FILTERS: List[Pattern] = [
    re.compile(r"Dokument bereits importiert"),
]
ALWAYS_DOSSIER_IMPORT_FILTERS: List[Pattern] = [
    re.compile(r"update-dossier"),
]


def parse_arguments():
    """
    Parse command line arguments for the script.

    Returns:
        argparse.Namespace: An object containing the parsed arguments:
            - input_paths (List[str]): One or more paths to migration report directories
            - output (str): Target directory for consolidated Excel files

    Raises:
        SystemExit: If required arguments are missing or invalid
    """
    parser = argparse.ArgumentParser(
        description="Consolidate migration reports into Excel files by municipality and segment.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process a single migration directory
  python consolidate_migration_reports.py /path/to/migration -o /path/to/output

  # Process multiple migration directories
  python consolidate_migration_reports.py /path/to/migration1 /path/to/migration2 -o /path/to/output
        """,
    )
    parser.add_argument(
        "input_paths",
        nargs="+",
        help="One or more paths to migration report directories",
    )
    parser.add_argument(
        "-o",
        "--output",
        required=True,
        help="Target directory for consolidated Excel files",
    )
    return parser.parse_args()


def create_directory(path: Path):
    """Create directory if it doesn't exist."""
    try:
        if not path.exists():
            path.mkdir()
            logger.info(f"Created directory: {path}")
    except Exception as e:
        logger.error(f"Failed to create directory {path}: {e}")
        raise


def extract_timestamp_from_dirname(dirname: str) -> Optional[datetime.datetime]:
    """
    Extract timestamp from directory name with format yyyy-mm-dd_HH-MM-SS_env[_suffix].

    Args:
        dirname: The directory name to extract timestamp from

    Returns:
        Datetime object if timestamp could be extracted, None otherwise
    """
    # Pattern to match yyyy-mm-dd_HH-MM-SS at the beginning of the string
    pattern = r"^(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})_"
    match = re.match(pattern, dirname)

    if match:
        date_str, time_str = match.groups()
        # Replace hyphens with colons in time part
        time_str = time_str.replace("-", ":")
        # Combine date and time
        datetime_str = f"{date_str} {time_str}"

        try:
            # Parse the datetime string
            return datetime.datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
        except ValueError as e:
            logger.warning(
                f"Failed to parse timestamp from directory name '{dirname}': {e}"
            )
            return None

    return None


def extract_municipality_name(filename: str) -> str:
    """
    Extract municipality name from filename, handling cases where municipality names contain hyphens.

    Args:
        filename: The filename (without extension) to extract municipality from

    Returns:
        The extracted municipality name
    """
    # Handle municipality names that may contain hyphens
    # Format is: municipality[-municipality][-mm_dd_HH_MM]
    # The timestamp part always has a specific format with underscores
    if "-" in filename and re.search(r"-\d{2}_\d{2}_\d{2}_\d{2}$", filename):
        # If there's a timestamp pattern at the end, extract everything before it
        return re.sub(r"-\d{2}_\d{2}_\d{2}_\d{2}$", "", filename)
    else:
        return filename


def extract_timestamp_from_csv_filename(filename: str) -> Optional[datetime.datetime]:
    """
    Extract timestamp from CSV filename with format ending in -mm_dd_HH_MM.

    Args:
        filename: The filename (without extension) to extract timestamp from

    Returns:
        Datetime object if timestamp could be extracted, None otherwise
    """
    # Pattern to match -mm_dd_HH_MM at the end of the string
    pattern = r"-(\d{2})_(\d{2})_(\d{2})_(\d{2})$"
    match = re.search(pattern, filename)

    if match:
        month, day, hour, minute = match.groups()
        # Since we don't have year in the filename, use the current year
        # or a reference year - this should work for ordering within the same migration
        current_year = datetime.datetime.now().year

        try:
            # Parse the datetime
            return datetime.datetime(
                current_year, int(month), int(day), int(hour), int(minute)
            )
        except ValueError as e:
            logger.warning(
                f"Failed to parse timestamp from CSV filename '{filename}': {e}"
            )
            return None

    return None


def sort_csv_files_by_timestamp(csv_files: List[Path]) -> List[Path]:
    """
    Sort CSV files by timestamp extracted from filenames.

    Files with valid timestamps come first (sorted by timestamp oldest first),
    followed by files without timestamps (in their original order).

    Args:
        csv_files: List of CSV file paths

    Returns:
        Sorted list of CSV file paths
    """
    files_with_timestamps = []
    files_without_timestamps = []

    for csv_file in csv_files:
        filename = csv_file.stem
        timestamp = extract_timestamp_from_csv_filename(filename)

        if timestamp:
            files_with_timestamps.append((csv_file, timestamp))
        else:
            files_without_timestamps.append(csv_file)

    # Sort files with timestamps by timestamp (oldest first)
    sorted_files_with_timestamps = [
        file for file, _ in sorted(files_with_timestamps, key=lambda x: x[1])
    ]

    # Combine sorted files with timestamps and files without timestamps
    sorted_files = sorted_files_with_timestamps + files_without_timestamps

    return sorted_files


def process_csv_file(
    csv_file: Path, segment_name: str, report_type: str, result_dict: Dict
) -> bool:
    """
    Process a single CSV file and add it to the result dictionary.

    Args:
        csv_file: Path to the CSV file
        segment_name: Name of the segment
        report_type: Type of report (document_export, document_import, dossier_import)
        result_dict: Dictionary to add the file to

    Returns:
        True if file was processed successfully, False otherwise
    """
    try:
        # Extract municipality name from filename
        filename = csv_file.stem
        municipality = extract_municipality_name(filename)

        if municipality not in result_dict[segment_name][report_type]:
            result_dict[segment_name][report_type][municipality] = []

        result_dict[segment_name][report_type][municipality].append(csv_file)
        return True
    except Exception as e:
        logger.error(f"Error processing CSV file {csv_file}: {e}")
        return False


def process_report_type(
    report_dir: Path, segment_name: str, report_type: str, result_dict: Dict
) -> int:
    """
    Process all CSV files in a report type directory.

    Args:
        report_dir: Path to the report type directory
        segment_name: Name of the segment
        report_type: Type of report (document_export, document_import, dossier_import)
        result_dict: Dictionary to add the files to

    Returns:
        Number of CSV files processed
    """
    try:
        if not report_dir.exists() or not report_dir.is_dir():
            return 0

        # Find CSV files
        csv_count = 0
        for csv_file in report_dir.glob("*.csv"):
            if process_csv_file(csv_file, segment_name, report_type, result_dict):
                csv_count += 1

        logger.info(f"Found {csv_count} CSV files in {report_dir}")
        return csv_count
    except Exception as e:
        logger.error(
            f"Error processing report type directory {report_type} in {report_dir.parent}: {e}"
        )
        return 0


def process_segment(segment_dir: Path, result_dict: Dict) -> None:
    """
    Process a segment directory and all its report type subdirectories.

    Args:
        segment_dir: Path to the segment directory
        result_dict: Dictionary to add the files to
    """
    if not segment_dir.is_dir():
        return

    segment_name = segment_dir.name
    if segment_name not in result_dict:
        result_dict[segment_name] = {
            "document_export": {},
            "document_import": {},
            "dossier_import": {},
        }

    # Check for report type directories
    for report_type in ["document_export", "document_import", "dossier_import"]:
        report_dir = segment_dir / report_type
        process_report_type(report_dir, segment_name, report_type, result_dict)


def process_input_path(input_path: str, result_dict: Dict) -> None:
    """
    Process an input path and all its segment subdirectories.

    Args:
        input_path: Path to the input directory
        result_dict: Dictionary to add the files to
    """
    try:
        # Convert to absolute path if it's relative
        input_dir = Path(input_path).resolve()
        if not input_dir.exists() or not input_dir.is_dir():
            logger.warning(
                f"Input path {input_path} does not exist or is not a directory. Skipping."
            )
            return

        # Find segment directories (1-4 subdirectories)
        for segment_dir in input_dir.iterdir():
            process_segment(segment_dir, result_dict)
    except Exception as e:
        logger.error(f"Error processing input path {input_path}: {e}")


def log_summary(result_dict: Dict) -> Tuple[int, int]:
    """
    Log a summary of the files found.

    Args:
        result_dict: Dictionary containing the found files

    Returns:
        Tuple of (total_municipalities, total_files)
    """
    total_municipalities = set()
    total_files = 0
    for segment, report_types in result_dict.items():
        for report_type, municipalities in report_types.items():
            for municipality, files in municipalities.items():
                total_municipalities.add(municipality)
                total_files += len(files)

    logger.info(
        f"Found a total of {total_files} CSV files for {len(total_municipalities)} municipalities across {len(result_dict)} segments"
    )
    return len(total_municipalities), total_files


def sort_input_paths_by_timestamp(input_paths: List[str]) -> List[str]:
    """
    Sort input paths by timestamp extracted from directory names.

    Directories with valid timestamps come first (sorted by timestamp),
    followed by directories without timestamps (in their original order).

    Args:
        input_paths: List of input directory paths

    Returns:
        Sorted list of input paths
    """
    paths_with_timestamps = []
    paths_without_timestamps = []

    for path in input_paths:
        # Get the directory name from the path
        dirname = Path(path).name
        timestamp = extract_timestamp_from_dirname(dirname)

        if timestamp:
            paths_with_timestamps.append((path, timestamp))
        else:
            paths_without_timestamps.append(path)

    # Sort paths with timestamps by timestamp (oldest first)
    sorted_paths_with_timestamps = [
        path for path, _ in sorted(paths_with_timestamps, key=lambda x: x[1])
    ]

    # Combine sorted paths with timestamps and paths without timestamps
    sorted_paths = sorted_paths_with_timestamps + paths_without_timestamps

    if sorted_paths != input_paths:
        logger.info("Input directories have been sorted by timestamp (oldest first)")
        for i, path in enumerate(sorted_paths):
            dirname = Path(path).name
            timestamp = extract_timestamp_from_dirname(dirname)
            timestamp_str = (
                timestamp.strftime("%Y-%m-%d %H:%M:%S") if timestamp else "No timestamp"
            )
            logger.info(f"  {i + 1}. {dirname} ({timestamp_str})")

    return sorted_paths


def find_csv_files(
    input_paths: List[str],
) -> Dict[str, Dict[str, Dict[str, List[Path]]]]:
    """
    Find all CSV files in the input paths and organize them by segment, report type, and municipality.

    Input paths are sorted by timestamp (oldest first) before processing.

    Returns:
        Dict with structure: {segment: {report_type: {municipality: [csv_files]}}}
    """
    # Main function logic
    result = {}

    try:
        # Sort input paths by timestamp
        sorted_input_paths = sort_input_paths_by_timestamp(input_paths)

        # Process each input path
        for input_path in sorted_input_paths:
            process_input_path(input_path, result)

        # Log summary of found files
        log_summary(result)
    except Exception as e:
        logger.error(f"Error in find_csv_files: {e}")
        raise

    return result


def should_filter_row(row_text: str, filters: List[Pattern]) -> bool:
    """
    Check if a row should be filtered out based on regex patterns.

    Args:
        row_text: Single-line representation of the row
        filters: List of regex patterns to match against

    Returns:
        True if the row should be filtered out, False otherwise
    """
    for pattern in filters:
        if pattern.search(row_text):
            return True
    return False


def get_filters_for_report_type(
    report_type: str, first_time: bool = True
) -> List[Pattern]:
    """
    Select the appropriate filter based on report type and whether it's the first time processing.

    Args:
        report_type: Type of report (document_export, document_import, or dossier_import)
        first_time: Whether this is the first time processing (oldest directory)

    Returns:
        List of regex patterns to use as filters
    """
    if first_time:
        # Filters for the oldest directory (first time processing)
        if report_type == "document_export":
            first_time_filters = FIRST_TIME_DOCUMENT_EXPORT_FILTERS
        elif report_type == "document_import":
            first_time_filters = FIRST_TIME_DOCUMENT_IMPORT_FILTERS
        elif report_type == "dossier_import":
            first_time_filters = FIRST_TIME_DOSSIER_IMPORT_FILTERS
        else:
            logger.warning(
                f"Unknown report type: {report_type}, no filters will be applied"
            )
            first_time_filters = []
    else:
        # Empty list for non-first time processing
        first_time_filters = []

    # Always applied filters
    if report_type == "document_export":
        always_filters = ALWAYS_DOCUMENT_EXPORT_FILTERS
    elif report_type == "document_import":
        always_filters = ALWAYS_DOCUMENT_IMPORT_FILTERS
    elif report_type == "dossier_import":
        always_filters = ALWAYS_DOSSIER_IMPORT_FILTERS
    else:
        always_filters = []

    # Combine both filter sets
    return first_time_filters + always_filters


def prepare_row_for_filtering(row: List[str]) -> str:
    """
    Prepare a row for filtering by replacing newlines with spaces and joining cells.

    Args:
        row: List of cell values

    Returns:
        Single string representation of the row for filtering
    """
    row_for_filtering = [cell.replace("\n", " ").replace("\r", " ") for cell in row]
    return " ".join(row_for_filtering)


def process_csv_row(
    row: List[str], filters: List[Pattern], csv_file: Path
) -> Tuple[bool, bool]:
    """
    Process a single CSV row and determine if it should be kept or filtered.

    Args:
        row: The CSV row to process
        filters: List of regex patterns to match against
        csv_file: Path to the CSV file (for error logging)

    Returns:
        Tuple of (should_keep, is_error)
    """
    try:
        row_text = prepare_row_for_filtering(row)
        should_keep = not should_filter_row(row_text, filters)
        return should_keep, False
    except Exception as e:
        logger.error(f"Error processing row in {csv_file}: {e}")
        return False, True


def handle_csv_headers(
    csv_reader: csv.reader, is_first_file: bool, csv_file: Path
) -> Tuple[Optional[List[str]], bool]:
    """
    Handle CSV headers based on whether this is the first file or not.

    Args:
        csv_reader: CSV reader object
        is_first_file: Whether this is the first file being processed
        csv_file: Path to the CSV file (for error logging)

    Returns:
        Tuple of (headers, should_continue)
        - headers: List of header values if this is the first file, None otherwise
        - should_continue: Whether processing should continue
    """
    try:
        headers = next(csv_reader)
        if is_first_file:
            logger.debug(f"Headers from first file: {headers}")
            return headers, True
        return None, True
    except StopIteration:
        logger.warning(f"CSV file {csv_file} is empty")
        return None, False


def process_csv_file_content(
    csv_file: Path, is_first_file: bool, filters: List[Pattern], report_type: str = None
) -> Tuple[List[List[str]], int, int]:
    """
    Process the content of a single CSV file.

    Args:
        csv_file: Path to the CSV file
        is_first_file: Whether this is the first file being processed
        filters: List of regex patterns to match against
        report_type: Type of report (document_export, document_import, or dossier_import)

    Returns:
        Tuple of (data, rows_processed, rows_filtered)
        - data: List of rows from the CSV file
        - rows_processed: Number of rows processed
        - rows_filtered: Number of rows filtered out
    """
    data = []
    rows_processed = 0
    rows_filtered = 0

    try:
        with open(csv_file, "r", encoding="utf-8") as f:
            try:
                csv_reader = csv.reader(f)

                # Handle headers
                headers, should_continue = handle_csv_headers(
                    csv_reader, is_first_file, csv_file
                )
                if not should_continue:
                    return data, rows_processed, rows_filtered

                if is_first_file and headers:
                    data.append(headers)

                # Process rows
                file_rows = 0
                file_filtered = 0
                for row in csv_reader:
                    rows_processed += 1
                    file_rows += 1

                    should_keep, is_error = process_csv_row(row, filters, csv_file)
                    if should_keep:
                        # For document_export files, replace "File exists already, skipping upload" with empty string in last column
                        if report_type == "document_export" and len(row) > 0:
                            if row[-1] == "File exists already, skipping upload":
                                row[-1] = ""
                                logger.info(
                                    "Replaced 'File exists already, skipping upload' with empty string in row"
                                )
                        data.append(row)
                    else:
                        rows_filtered += 1
                        file_filtered += 1

                logger.info(
                    f"Processed {file_rows} rows from {csv_file}, filtered out {file_filtered} rows"
                )
            except csv.Error as e:
                logger.error(f"CSV error in file {csv_file}: {e}")
    except Exception as e:
        logger.error(f"Error opening file {csv_file}: {e}")

    return data, rows_processed, rows_filtered


def find_key_column_indices(headers: List[str]) -> Optional[Tuple[int, int, int]]:
    """
    Find the indices of the key columns for document_export de-duplication.

    Args:
        headers: List of column headers

    Returns:
        Tuple of (gemeinde_idx, laufnummer_idx, gesuch_id_idx) if all columns are found, None otherwise
    """
    try:
        # Find indices of the key columns
        gemeinde_idx = headers.index("Gemeinde")
        laufnummer_idx = headers.index("Laufnummer")
        gesuch_id_idx = headers.index("Gesuch-ID")
        key_column_indices = (gemeinde_idx, laufnummer_idx, gesuch_id_idx)
        logger.info(
            f"Found key columns for document_export de-duplication: {key_column_indices}"
        )
        return key_column_indices
    except ValueError as e:
        logger.warning(
            f"Could not find all key columns for document_export de-duplication: {e}"
        )
        return None


def process_document_export_row(
    row: List[str],
    key_column_indices: Tuple[int, int, int],
    unique_records: Dict[Tuple[str, str, str], bool],
) -> Tuple[bool, Tuple[str, str, str]]:
    """
    Process a document_export row for de-duplication.

    Args:
        row: The row to process
        key_column_indices: Tuple of (gemeinde_idx, laufnummer_idx, gesuch_id_idx)
        unique_records: Dictionary of already seen unique records

    Returns:
        Tuple of (should_add, record_key)
        - should_add: Whether the row should be added to the result
        - record_key: The unique key for this record
    """
    # Skip rows that don't have enough columns
    if len(row) <= max(key_column_indices):
        logger.warning(f"Row has fewer columns than expected, skipping: {row}")
        return False, ("", "", "")

    # Create a unique key from the three key columns
    record_key = (
        row[key_column_indices[0]],  # Gemeinde
        row[key_column_indices[1]],  # Laufnummer
        row[key_column_indices[2]],  # Gesuch-ID
    )

    # If this record hasn't been seen before, it should be added
    should_add = record_key not in unique_records

    return should_add, record_key


def deduplicate_document_export_data(
    file_data: List[List[str]],
    is_first_file: bool,
    key_column_indices: Optional[Tuple[int, int, int]],
    unique_records: Dict[Tuple[str, str, str], bool],
    csv_file: Path,
) -> Tuple[List[List[str]], int]:
    """
    De-duplicate document_export data based on key columns.

    Args:
        file_data: The data from the CSV file
        is_first_file: Whether this is the first file being processed
        key_column_indices: Tuple of (gemeinde_idx, dateiname_idx, gesuch_id_idx)
        unique_records: Dictionary of already seen unique records
        csv_file: Path to the CSV file (for logging)

    Returns:
        Tuple of (deduplicated_data, duplicates_found)
        - deduplicated_data: The de-duplicated data
        - duplicates_found: Number of duplicates found
    """
    if not file_data:
        return [], 0

    # If we don't have key column indices, just return the data as is
    if not key_column_indices:
        return file_data, 0

    deduplicated_data = []
    duplicates_found = 0

    # Add header row if this is the first file
    if is_first_file:
        deduplicated_data.append(file_data[0])

    # Process data rows (skip header if this is the first file)
    for row in file_data[1:] if is_first_file else file_data:
        should_add, record_key = process_document_export_row(
            row, key_column_indices, unique_records
        )

        if should_add:
            unique_records[record_key] = True
            deduplicated_data.append(row)
        else:
            duplicates_found += 1

    if duplicates_found > 0:
        logger.info(f"Skipped {duplicates_found} duplicate records in {csv_file}")

    return deduplicated_data, duplicates_found


def determine_filters_to_apply(
    csv_file: Path, parent_dir: str, report_type: str, oldest_import_dir: Optional[str]
) -> List[Pattern]:
    """
    Determine which filters to apply based on chronological order and report type.

    Args:
        csv_file: Path to the CSV file (for logging)
        parent_dir: Parent directory of the CSV file
        report_type: Type of report (document_export, document_import, or dossier_import)
        oldest_import_dir: Path to the oldest import directory (chronologically)

    Returns:
        List of regex patterns to use as filters
    """
    if parent_dir == oldest_import_dir:
        # For the oldest import directory, apply both first-time and always filters
        filters = get_filters_for_report_type(report_type, first_time=True)
        logger.info(
            f"Applying first-time and always filters to {csv_file} (oldest import directory)"
        )
    else:
        # For other directories, apply only the always filters
        filters = get_filters_for_report_type(report_type, first_time=False)
        if filters:
            logger.info(f"Applying always filters to {csv_file}")
        else:
            logger.info(f"No filters applied to {csv_file} (no always filters defined)")

    return filters


def log_processing_summary(
    report_type: str,
    total_rows_processed: int,
    total_rows_filtered: int,
    all_data: List[List[str]],
    key_column_indices: Optional[Tuple[int, int, int]] = None,
    total_duplicates: int = 0,
) -> None:
    """
    Log a summary of the CSV processing.

    Args:
        report_type: Type of report (document_export, document_import, or dossier_import)
        total_rows_processed: Total number of rows processed
        total_rows_filtered: Total number of rows filtered out
        all_data: The final combined data
        key_column_indices: Tuple of (gemeinde_idx, dateiname_idx, gesuch_id_idx) for document_export
        total_duplicates: Total number of duplicates found for document_export
    """
    if report_type == "document_export" and key_column_indices:
        logger.info(
            f"Total rows processed: {total_rows_processed}, filtered out: {total_rows_filtered}, "
            f"de-duplicated: {total_duplicates}, "
            f"remaining: {len(all_data) - 1 if all_data else 0}"
        )
    else:
        logger.info(
            f"Total rows processed: {total_rows_processed}, filtered out: {total_rows_filtered}, "
            f"remaining: {len(all_data) - 1 if all_data else 0}"
        )


def process_single_csv_file(
    csv_file: Path,
    report_type: str,
    is_first_file: bool,
    oldest_import_dir: Optional[str],
    unique_records: Dict[Tuple[str, str, str], bool],
    key_column_indices: Optional[Tuple[int, int, int]],
) -> Tuple[List[List[str]], int, int, int, Optional[Tuple[int, int, int]]]:
    """
    Process a single CSV file and handle de-duplication if needed.

    Args:
        csv_file: Path to the CSV file
        report_type: Type of report (document_export, document_import, or dossier_import)
        is_first_file: Whether this is the first file being processed
        oldest_import_dir: Path to the oldest import directory (chronologically)
        unique_records: Dictionary of already seen unique records for document_export
        key_column_indices: Tuple of (gemeinde_idx, dateiname_idx, gesuch_id_idx) for document_export

    Returns:
        Tuple of (processed_data, rows_processed, rows_filtered, duplicates_found, updated_key_column_indices)
    """
    logger.info(f"Processing {csv_file}")

    # Get the parent directory of the CSV file (the import directory)
    parent_dir = str(csv_file.resolve().parent.parent.parent)

    # Determine which filters to apply based on chronological order and report type
    filters = determine_filters_to_apply(
        csv_file, parent_dir, report_type, oldest_import_dir
    )

    # Process the CSV file content
    file_data, rows_processed, rows_filtered = process_csv_file_content(
        csv_file,
        is_first_file=is_first_file,
        filters=filters,
        report_type=report_type,
    )

    # For document_export, de-duplicate based on key columns
    duplicates_found = 0
    processed_data = []

    if report_type == "document_export" and file_data:
        # If this is the first file with data, get the indices of the key columns
        if is_first_file or not key_column_indices:
            headers = file_data[0]
            key_column_indices = find_key_column_indices(headers)

        # De-duplicate the data
        deduplicated_data, duplicates_found = deduplicate_document_export_data(
            file_data,
            is_first_file=is_first_file,
            key_column_indices=key_column_indices,
            unique_records=unique_records,
            csv_file=csv_file,
        )

        processed_data = deduplicated_data
    else:
        # For other report types, just add all rows
        processed_data = file_data

    return (
        processed_data,
        rows_processed,
        rows_filtered,
        duplicates_found,
        key_column_indices,
    )


def is_empty_value(value: str) -> bool:
    """
    Check if a value is considered empty.

    Args:
        value: The value to check

    Returns:
        True if the value is None, "", "-", or " - ", False otherwise
    """
    if value is None:
        return True
    if isinstance(value, str):
        return value in ("", "-", " - ")
    return False


def initialize_central_data_structure() -> Dict[str, Dict]:
    """
    Initialize the centralized data structure for all three report types.

    Structure: {
        "document_export": {segment: {gemeinde: {gesuch_id: {laufnummer: [rows]}}}},
        "document_import": {segment: {gemeinde: {gesuch_id: {dateiname: [rows]}}}},
        "dossier_import": {segment: {gemeinde: {gesuch_id: [rows]}}}
    }

    Returns:
        Dict with initialized nested structure for each report type
    """
    return {
        "document_export": {},
        "document_import": {},
        "dossier_import": {},
    }


def get_nested_value(nested_dict: Dict, keys: List[str], default=None):
    """
    Safely get a value from a nested dictionary.

    Args:
        nested_dict: The nested dictionary
        keys: List of keys to traverse
        default: Default value if path doesn't exist

    Returns:
        The value at the path, or default if not found
    """
    current = nested_dict
    for key in keys:
        if not isinstance(current, dict) or key not in current:
            return default
        current = current[key]
    return current


def set_nested_value(nested_dict: Dict, keys: List[str], value):
    """
    Set a value in a nested dictionary, creating intermediate dicts as needed.

    Args:
        nested_dict: The nested dictionary
        keys: List of keys to traverse
        value: Value to set at the path
    """
    current = nested_dict
    for key in keys[:-1]:
        if key not in current:
            current[key] = {}
        current = current[key]
    current[keys[-1]] = value


def add_row_to_central_structure(  # noqa: C901
    central_data: Dict,
    report_type: str,
    segment: str,
    row: List[str],
    headers: List[str],
) -> bool:
    """
    Add a parsed row to the central data structure.

    Args:
        central_data: The central data structure
        report_type: Type of report (document_export, document_import, dossier_import)
        segment: Segment name
        row: The parsed row data
        headers: Column headers

    Returns:
        True if row was added successfully, False otherwise
    """
    try:
        # Get key column names for this report type
        key_columns, _ = get_key_columns_for_report_type(report_type)
        if not key_columns:
            return False

        # Extract key values from row
        gemeinde_col = key_columns[1]  # Skip "Segment" which is from parent path
        gesuch_id_col = key_columns[2]

        try:
            gemeinde_idx = headers.index(gemeinde_col)
            gesuch_id_idx = headers.index(gesuch_id_col)
        except ValueError as e:
            logger.warning(f"Could not find key columns in headers: {e}")
            return False

        gemeinde = row[gemeinde_idx] if gemeinde_idx < len(row) else ""
        gesuch_id = row[gesuch_id_idx] if gesuch_id_idx < len(row) else ""

        # For document_* reports, also extract the 4th key column (Dateiname or Laufnummer depending on type)
        if report_type in ["document_export", "document_import"]:
            dateiname_col = key_columns[3]
            try:
                dateiname_idx = headers.index(dateiname_col)
                dateiname = row[dateiname_idx] if dateiname_idx < len(row) else ""
            except ValueError:
                logger.warning(
                    f"Could not find column '{dateiname_col}' in headers for {report_type}"
                )
                return False

            # Build path: segment -> gemeinde -> gesuch_id -> dateiname -> list
            if segment not in central_data[report_type]:
                central_data[report_type][segment] = {}
            if gemeinde not in central_data[report_type][segment]:
                central_data[report_type][segment][gemeinde] = {}
            if gesuch_id not in central_data[report_type][segment][gemeinde]:
                central_data[report_type][segment][gemeinde][gesuch_id] = {}
            if dateiname not in central_data[report_type][segment][gemeinde][gesuch_id]:
                central_data[report_type][segment][gemeinde][gesuch_id][dateiname] = []

            central_data[report_type][segment][gemeinde][gesuch_id][dateiname].append(
                row
            )
        else:
            # dossier_import: segment -> gemeinde -> gesuch_id -> list
            if segment not in central_data[report_type]:
                central_data[report_type][segment] = {}
            if gemeinde not in central_data[report_type][segment]:
                central_data[report_type][segment][gemeinde] = {}
            if gesuch_id not in central_data[report_type][segment][gemeinde]:
                central_data[report_type][segment][gemeinde][gesuch_id] = []

            central_data[report_type][segment][gemeinde][gesuch_id].append(row)

        return True
    except Exception as e:
        logger.error(f"Error adding row to central structure: {e}")
        return False


def parse_all_csv_files_to_central_structure(  # noqa: C901
    input_paths: List[str],
) -> Tuple[Dict, Dict[str, List[str]]]:
    """
    Parse all CSV files from all input directories chronologically into the central data structure.

    Args:
        input_paths: List of input directory paths

    Returns:
        Tuple of (central_data, headers_by_report_type)
        - central_data: The populated central data structure
        - headers_by_report_type: Dict mapping report_type to its headers
    """
    central_data = initialize_central_data_structure()
    headers_by_report_type = {}

    # Sort input paths chronologically
    sorted_input_paths = sort_input_paths_by_timestamp(input_paths)

    # Determine oldest import directory for filter selection
    oldest_import_dir = None
    if sorted_input_paths:
        oldest_import_dir = str(Path(sorted_input_paths[0]).resolve())
        logger.info(f"Oldest import directory: {oldest_import_dir}")

    # Process each input directory chronologically
    for input_path in sorted_input_paths:
        input_dir = Path(input_path).resolve()
        if not input_dir.exists() or not input_dir.is_dir():
            logger.warning(f"Skipping non-existent directory: {input_path}")
            continue

        logger.info(f"Processing input directory: {input_dir}")

        # Process each segment directory
        for segment_dir in sorted(input_dir.iterdir()):
            if not segment_dir.is_dir():
                continue

            segment_name = segment_dir.name
            logger.info(f"  Processing segment: {segment_name}")

            # Process each report type directory
            for report_type in ["document_export", "document_import", "dossier_import"]:
                report_dir = segment_dir / report_type
                if not report_dir.exists() or not report_dir.is_dir():
                    continue

                logger.info(f"    Processing report type: {report_type}")

                # Get all CSV files and sort them chronologically
                csv_files = list(report_dir.glob("*.csv"))
                sorted_csv_files = sort_csv_files_by_timestamp(csv_files)

                # Determine filters for this directory and report type
                parent_dir = str(input_dir)
                filters = determine_filters_to_apply(
                    report_dir / "dummy.csv",  # Just for logging
                    parent_dir,
                    report_type,
                    oldest_import_dir,
                )

                # Process each CSV file
                for csv_file in sorted_csv_files:
                    logger.info(f"      Processing CSV: {csv_file.name}")

                    try:
                        with open(csv_file, "r", encoding="utf-8") as f:
                            csv_reader = csv.reader(f)

                            # Read headers
                            try:
                                headers = next(csv_reader)
                            except StopIteration:
                                logger.warning(f"Empty CSV file: {csv_file}")
                                continue

                            # Store headers for this report type (first occurrence)
                            if report_type not in headers_by_report_type:
                                headers_by_report_type[report_type] = headers
                                logger.info(
                                    f"Stored headers for {report_type}: {headers}"
                                )

                            # Process each row
                            rows_added = 0
                            rows_filtered = 0
                            for row in csv_reader:
                                # Apply filters
                                row_text = prepare_row_for_filtering(row)
                                if should_filter_row(row_text, filters):
                                    rows_filtered += 1
                                    continue

                                # Apply substitutions for document_export
                                if report_type == "document_export" and len(row) > 0:
                                    if (
                                        row[-1]
                                        == "File exists already, skipping upload"
                                    ):
                                        row[-1] = ""

                                # Add row to central structure
                                if add_row_to_central_structure(
                                    central_data,
                                    report_type,
                                    segment_name,
                                    row,
                                    headers,
                                ):
                                    rows_added += 1

                            logger.info(
                                f"        Added {rows_added} rows, filtered {rows_filtered} rows"
                            )

                    except Exception as e:
                        logger.error(f"Error processing CSV file {csv_file}: {e}")
                        continue

    return central_data, headers_by_report_type


def flatten_central_data_to_rows(
    central_data: Dict, report_type: str, headers: List[str]
) -> List[List[str]]:
    """
    Flatten the central data structure to a list of rows for a report type.

    Args:
        central_data: The central data structure
        report_type: Type of report
        headers: Column headers

    Returns:
        List of rows including header
    """
    result = [headers]

    if report_type not in central_data:
        return result

    # Traverse the nested structure
    for segment in sorted(central_data[report_type].keys()):
        for gemeinde in sorted(central_data[report_type][segment].keys()):
            for gesuch_id in sorted(
                central_data[report_type][segment][gemeinde].keys()
            ):
                if report_type in ["document_export", "document_import"]:
                    # Has dateiname level
                    for dateiname in sorted(
                        central_data[report_type][segment][gemeinde][gesuch_id].keys()
                    ):
                        rows = central_data[report_type][segment][gemeinde][gesuch_id][
                            dateiname
                        ]
                        result.extend(rows)
                else:
                    # dossier_import: no dateiname level
                    rows = central_data[report_type][segment][gemeinde][gesuch_id]
                    result.extend(rows)

    return result


def get_consolidated_row_from_list(
    rows: List[List[str]], error_column_idx: int
) -> Optional[List[str]]:
    """
    Get the consolidated row from a list: first row without error, or last row if none exist.

    Args:
        rows: List of rows for the same key
        error_column_idx: Index of the error column

    Returns:
        The selected row, or None if list is empty
    """
    if not rows:
        return None

    # Find first row without error
    for row in rows:
        if not has_error(row, error_column_idx):
            return row

    # No row without error, return last row
    return rows[-1]


def has_empty_key_component(key_tuple: Tuple[str, ...]) -> bool:
    """
    Check if any component of the key is empty.

    Args:
        key_tuple: Tuple of key components

    Returns:
        True if any component is empty
    """
    return any(is_empty_value(v) for v in key_tuple)


def is_more_specific_key(key1: Tuple[str, ...], key2: Tuple[str, ...]) -> bool:
    """
    Check if key1 is more specific than key2 (has same prefix but more non-empty values).

    Args:
        key1: First key tuple
        key2: Second key tuple

    Returns:
        True if key1 is more specific than key2
    """
    if len(key1) != len(key2):
        return False

    # Count non-empty values
    non_empty1 = sum(1 for v in key1 if not is_empty_value(v))
    non_empty2 = sum(1 for v in key2 if not is_empty_value(v))

    if non_empty1 <= non_empty2:
        return False

    # Check if key1 extends key2 (all non-empty values in key2 match key1)
    for v1, v2 in zip(key1, key2):
        if not is_empty_value(v2) and v1 != v2:
            return False

    return True


def flatten_central_data_to_rows_consolidated(  # noqa: C901
    central_data: Dict, report_type: str, headers: List[str]
) -> List[List[str]]:
    """
    Flatten the central data structure to consolidated rows for a report type.

    For each unique key:
    - Select first row without error, or last row if none exist
    - Exclude rows with empty key components if more specific keys exist

    Args:
        central_data: The central data structure
        report_type: Type of report
        headers: Column headers

    Returns:
        List of rows including header
    """
    result = [headers]

    if report_type not in central_data:
        return result

    # Get error column index
    _, error_column_name = get_key_columns_for_report_type(report_type)
    try:
        error_column_idx = headers.index(error_column_name)
    except ValueError:
        logger.warning(
            f"Error column '{error_column_name}' not found for {report_type}"
        )
        error_column_idx = -1

    # Collect all keys and their consolidated rows
    keys_and_rows = []

    # Traverse the nested structure
    for segment in sorted(central_data[report_type].keys()):
        for gemeinde in sorted(central_data[report_type][segment].keys()):
            for gesuch_id in sorted(
                central_data[report_type][segment][gemeinde].keys()
            ):
                if report_type in ["document_export", "document_import"]:
                    # Has dateiname level
                    for dateiname in sorted(
                        central_data[report_type][segment][gemeinde][gesuch_id].keys()
                    ):
                        rows = central_data[report_type][segment][gemeinde][gesuch_id][
                            dateiname
                        ]
                        key = (segment, gemeinde, gesuch_id, dateiname)
                        consolidated_row = get_consolidated_row_from_list(
                            rows, error_column_idx
                        )
                        if consolidated_row:
                            keys_and_rows.append((key, consolidated_row))
                else:
                    # dossier_import: no dateiname level
                    rows = central_data[report_type][segment][gemeinde][gesuch_id]
                    key = (segment, gemeinde, gesuch_id)
                    consolidated_row = get_consolidated_row_from_list(
                        rows, error_column_idx
                    )
                    if consolidated_row:
                        keys_and_rows.append((key, consolidated_row))

    # Filter out keys with empty components if more specific keys exist
    keys_to_keep = set()
    for i, (key1, _) in enumerate(keys_and_rows):
        should_keep = True
        if has_empty_key_component(key1):
            # Check if there's a more specific key
            for j, (key2, _) in enumerate(keys_and_rows):
                if i != j and is_more_specific_key(key2, key1):
                    should_keep = False
                    break
        if should_keep:
            keys_to_keep.add(i)

    # Add kept rows to result
    for i in keys_to_keep:
        _, row = keys_and_rows[i]
        result.append(row)

    return result


def get_sort_column_names_for_report_type(report_type: str) -> List[str]:
    """
    Get the column names used for sorting for a report type.

    Returns column names in order: Gemeinde, Gesuch-ID, Dateiname (if applicable), Import-Zeit
    Note: Segment is handled separately as it's from the key structure

    Args:
        report_type: Type of report

    Returns:
        List of column names for sorting
    """
    if report_type == "document_export":
        # Sort by Gemeinde, Gesuch-ID, then Laufnummer (asc), then Importzeit (older first)
        return ["Gemeinde", "Gesuch-ID", "Laufnummer", "Importzeit"]
    elif report_type == "document_import":
        # Sort by gemeinde, gesuch_id, then laufnummer_dok (asc), then importzeit (older first)
        return ["gemeinde", "gesuch_id", "laufnummer_dok", "importzeit"]
    elif report_type == "dossier_import":
        return ["Gemeinde", "Gesuch-ID", "Importzeit"]
    else:
        return []


def _get_sort_indices_or_warn(
    headers: List[str], report_type: str, sort_columns: List[str]
) -> Optional[List[int]]:
    """
    Resolve indices of required sort columns from headers.

    Returns None and logs a warning if any column is missing.
    """
    indices: List[int] = []
    for col_name in sort_columns:
        try:
            indices.append(headers.index(col_name))
        except ValueError:
            logger.warning(
                f"Sort column '{col_name}' not found in headers for {report_type}"
            )
            return None
    return indices


def _make_row_sort_key_factory(headers: List[str], sort_indices: List[int]):
    """
    Create a key function for sorting data rows according to the resolved indices.

    Numeric-aware for Laufnummer/laufnummer_dok and case-insensitive for text.
    """
    numeric_sort_columns = {"Laufnummer", "laufnummer_dok"}

    def num_val(v: str) -> int:
        if is_empty_value(v):
            return -1
        try:
            return int(v)
        except Exception:
            m = re.search(r"\d+", str(v))
            return int(m.group(0)) if m else -1

    def txt_val(v: str) -> str:
        if is_empty_value(v):
            return ""
        return v.lower() if isinstance(v, str) else str(v)

    def key_fn(row: List[str]):
        key_parts: List[object] = []
        for idx in sort_indices:
            header_name = headers[idx] if idx < len(headers) else ""
            val = row[idx] if idx < len(row) else ""
            if header_name in numeric_sort_columns:
                key_parts.append((num_val(val),))
            else:
                key_parts.append(txt_val(val))
        return tuple(key_parts)

    return key_fn


def sort_rows_by_keys(
    rows: List[List[str]], report_type: str, headers: List[str]
) -> List[List[str]]:
    """
    Sort rows by Gemeinde, Gesuch-ID and type-specific columns.

    Args:
        rows: List of rows including header
        report_type: Type of report
        headers: Column headers

    Returns:
        Sorted list of rows with header first
    """
    if len(rows) <= 1:
        return rows

    sort_columns = get_sort_column_names_for_report_type(report_type)
    if not sort_columns:
        return rows

    sort_indices = _get_sort_indices_or_warn(headers, report_type, sort_columns)
    if sort_indices is None:
        return rows

    header, data_rows = rows[0], rows[1:]
    key_fn = _make_row_sort_key_factory(headers, sort_indices)
    sorted_data_rows = sorted(data_rows, key=key_fn)
    return [header] + sorted_data_rows


def get_key_columns_for_report_type(report_type: str) -> Tuple[List[str], str]:
    """
    Get the key columns and error column for a report type.

    Args:
        report_type: Type of report (document_export, document_import, or dossier_import)

    Returns:
        Tuple of (key_columns, error_column)
    """
    if report_type == "document_export":
        # Use Laufnummer instead of Dateiname for the key hierarchy
        return (["Segment", "Gemeinde", "Gesuch-ID", "Laufnummer"], "Warnungen/Fehler")
    elif report_type == "document_import":
        return (["Segment", "gemeinde", "gesuch_id", "dateiname"], "fehler")
    elif report_type == "dossier_import":
        return (["Segment", "Gemeinde", "Gesuch-ID"], "Warnungen/Fehler")
    else:
        return ([], "")


def extract_compound_key(
    row: List[str], key_indices: List[int], segment: str
) -> Tuple[str, ...]:
    """
    Extract compound key from a row.

    Args:
        row: The row to extract key from
        key_indices: Indices of key columns (excluding Segment which is always first)
        segment: The segment name

    Returns:
        Tuple representing the compound key (Segment always first)
    """
    key_values = [segment]
    for idx in key_indices:
        if idx < len(row):
            key_values.append(row[idx])
        else:
            key_values.append("")
    return tuple(key_values)


def is_row_more_specific(key1: Tuple[str, ...], key2: Tuple[str, ...]) -> bool:
    """
    Check if key1 is more specific than key2.

    A key is more specific if it has the same non-empty prefix but more non-empty values.

    Args:
        key1: First compound key
        key2: Second compound key

    Returns:
        True if key1 is more specific than key2
    """
    # Both keys must have the same length
    if len(key1) != len(key2):
        return False

    # Check if key1 has more non-empty values
    non_empty1 = sum(1 for v in key1 if not is_empty_value(v))
    non_empty2 = sum(1 for v in key2 if not is_empty_value(v))

    if non_empty1 <= non_empty2:
        return False

    # Check if all non-empty values in key2 match key1 at the same positions
    for i, (v1, v2) in enumerate(zip(key1, key2)):
        if not is_empty_value(v2):
            if v1 != v2:
                return False

    return True


def has_error(row: List[str], error_column_idx: int) -> bool:
    """
    Check if a row has an error.

    Args:
        row: The row to check
        error_column_idx: Index of the error column

    Returns:
        True if the row has an error (non-empty error column)
    """
    if error_column_idx < 0 or error_column_idx >= len(row):
        return False
    return not is_empty_value(row[error_column_idx])


def build_specificity_index(key: Tuple[str, ...]) -> Tuple[int, Tuple[str, ...]]:
    """
    Build a specificity index for a compound key.

    Returns a tuple with:
    - Count of non-empty values
    - Normalized key (for matching prefixes)

    Args:
        key: Compound key tuple

    Returns:
        Tuple of (specificity_count, normalized_key)
    """
    specificity = sum(1 for v in key if not is_empty_value(v))
    # Create normalized key: keep non-empty values, mark positions of empty values
    normalized = tuple(v if not is_empty_value(v) else None for v in key)
    return (specificity, normalized)


def keys_match_prefix(key1_normalized: Tuple, key2_normalized: Tuple) -> bool:
    """
    Check if two normalized keys match on all non-None positions.

    Args:
        key1_normalized: First normalized key
        key2_normalized: Second normalized key

    Returns:
        True if they match on all non-None positions
    """
    if len(key1_normalized) != len(key2_normalized):
        return False

    for v1, v2 in zip(key1_normalized, key2_normalized):
        if v1 is not None and v2 is not None and v1 != v2:
            return False

    return True


def apply_error_based_replacement(  # noqa: C901
    all_data: List[List[str]], report_type: str, segment: str
) -> List[List[str]]:
    """
    Apply error-based replacement logic to the data using optimized hash-based approach.

    Newer entries with specific compound keys replace older entries with errors
    if the older entry is less specific (has empty/"-"/" - " values in key columns).

    Args:
        all_data: List of all rows including header
        report_type: Type of report
        segment: Segment name

    Returns:
        List of rows after applying error-based replacement
    """
    if not all_data or len(all_data) < 2:
        return all_data

    key_columns, error_column = get_key_columns_for_report_type(report_type)
    if not key_columns:
        return all_data

    headers = all_data[0]

    # Find indices of key columns (excluding Segment which we'll add)
    key_indices = []
    for col in key_columns[1:]:  # Skip "Segment" as it's from parent path
        try:
            key_indices.append(headers.index(col))
        except ValueError:
            logger.warning(f"Key column '{col}' not found in headers for {report_type}")
            return all_data

    # Find error column index
    try:
        error_column_idx = headers.index(error_column)
    except ValueError:
        logger.warning(
            f"Error column '{error_column}' not found in headers for {report_type}"
        )
        return all_data

    # Build optimized data structure:
    # - Dict mapping full compound key to (row_index, has_error, specificity, normalized_key)
    # - Group keys by specificity level for faster lookups
    rows_by_key = {}
    keys_by_specificity = {}  # specificity_level -> set of keys

    for i, row in enumerate(all_data[1:], start=1):  # Skip header
        key = extract_compound_key(row, key_indices, segment)
        row_has_error = has_error(row, error_column_idx)
        specificity, normalized_key = build_specificity_index(key)

        # Check if this row should replace any existing rows with lower specificity
        if not row_has_error:
            # This row has no error, check all less specific keys
            for lower_spec in range(specificity):
                if lower_spec in keys_by_specificity:
                    keys_to_check = list(keys_by_specificity[lower_spec])
                    for existing_key in keys_to_check:
                        if existing_key in rows_by_key:
                            _, existing_has_error, _, existing_normalized = rows_by_key[
                                existing_key
                            ]
                            # If existing has error and keys match on prefix, remove it
                            if existing_has_error and keys_match_prefix(
                                normalized_key, existing_normalized
                            ):
                                del rows_by_key[existing_key]
                                keys_by_specificity[lower_spec].discard(existing_key)

        # Check if current row should be kept
        should_add = True
        if row_has_error:
            # If current row has error, check if any more specific row already exists
            for higher_spec in range(specificity + 1, len(key) + 1):
                if higher_spec in keys_by_specificity:
                    for existing_key in keys_by_specificity[higher_spec]:
                        if existing_key in rows_by_key:
                            _, _, _, existing_normalized = rows_by_key[existing_key]
                            # If more specific key matches prefix, don't add current row
                            if keys_match_prefix(existing_normalized, normalized_key):
                                should_add = False
                                break
                    if not should_add:
                        break

        if should_add:
            rows_by_key[key] = (i, row_has_error, specificity, normalized_key)
            if specificity not in keys_by_specificity:
                keys_by_specificity[specificity] = set()
            keys_by_specificity[specificity].add(key)

    # Build final result preserving chronological order
    # Instead of iterating through all_data, extract indices from rows_by_key and sort
    result = [all_data[0]]  # Add header

    # Extract row indices from rows_by_key and sort them
    kept_indices = sorted(row_info[0] for row_info in rows_by_key.values())

    # Add rows in chronological order
    for idx in kept_indices:
        result.append(all_data[idx])

    removed_count = len(all_data) - 1 - len(kept_indices)
    if removed_count > 0:
        logger.info(
            f"Removed {removed_count} rows based on error replacement logic for {report_type} in segment {segment}"
        )

    return result


def process_csv_files(
    csv_files: List[Path],
    report_type: str,
    oldest_import_dir: Optional[str] = None,
    segment: str = "",
) -> List[List[str]]:
    """
    Process CSV files and return a list of rows with the combined data.

    Applies filtering based on chronological order and report type:
    - For files from the oldest import directory: applies all filters for the report type
    - For dossier_import files: always applies DOSSIER_IMPORT_FILTERS regardless of directory
    - For other report types from non-oldest directories: doesn't apply any filters

    For document_export files, de-duplicates rows based on the columns "Gemeinde", "Dateiname", and "Gesuch-ID".
    Chronologically earlier records are preserved when duplicates are found.

    Applies error-based replacement: newer entries with specific compound keys replace older entries
    with errors if the older entry is less specific.

    Args:
        csv_files: List of CSV files to process
        report_type: Type of report (document_export, document_import, or dossier_import)
        oldest_import_dir: Path to the oldest import directory (chronologically)
        segment: Segment name for error-based replacement logic

    Returns:
        List of rows with combined data from all CSV files
    """
    if not csv_files:
        logger.info(f"No CSV files to process for report type: {report_type}")
        return []

    try:
        all_data = []
        total_rows_processed = 0
        total_rows_filtered = 0
        total_duplicates = 0

        # For document_export, we need to track unique records by the three key columns
        unique_records = {}  # Dictionary to track unique records for document_export
        key_column_indices = (
            None  # Will store indices of the key columns for document_export
        )

        # Sort CSV files chronologically
        sorted_csv_files = sort_csv_files_by_timestamp(csv_files)

        for i, csv_file in enumerate(sorted_csv_files):
            # Process the CSV file
            (
                processed_data,
                rows_processed,
                rows_filtered,
                duplicates_found,
                updated_key_indices,
            ) = process_single_csv_file(
                csv_file=csv_file,
                report_type=report_type,
                is_first_file=(i == 0),
                oldest_import_dir=oldest_import_dir,
                unique_records=unique_records,
                key_column_indices=key_column_indices,
            )

            # Update key column indices if needed
            if updated_key_indices:
                key_column_indices = updated_key_indices

            # Add the processed data to the result
            if (
                i > 0
                and processed_data
                and len(processed_data) > 0
                and report_type != "document_export"
            ):
                # Skip header row for subsequent files (except for document_export which handles this in deduplicate_document_export_data)
                all_data.extend(processed_data[1:])
            else:
                all_data.extend(processed_data)

            # Update statistics
            total_rows_processed += rows_processed
            total_rows_filtered += rows_filtered
            total_duplicates += duplicates_found

        # Log summary
        log_processing_summary(
            report_type=report_type,
            total_rows_processed=total_rows_processed,
            total_rows_filtered=total_rows_filtered,
            all_data=all_data,
            key_column_indices=key_column_indices,
            total_duplicates=total_duplicates,
        )

        # Apply error-based replacement logic
        if segment:
            all_data = apply_error_based_replacement(all_data, report_type, segment)

        return all_data

    except Exception as e:
        logger.error(f"Error in process_csv_files for report type {report_type}: {e}")
        logger.error(traceback.format_exc())
        # Return empty list in case of error to allow processing to continue
        return []


def write_to_excel(data: List[List[str]], workbook: openpyxl.Workbook, sheet_name: str):
    """
    Write data to an Excel worksheet with adjusted column widths and autofilter.

    This function:
    - Writes the data to the worksheet
    - Adds an autofilter to the header row (first row)
    - Adjusts column widths based on content length
    - Handles multiline cells when calculating column widths

    Args:
        data: List of rows to write
        workbook: Excel workbook to write to
        sheet_name: Name of the worksheet to write to
    """
    try:
        if not data:
            # Create an empty worksheet
            logger.info(f"Creating empty worksheet '{sheet_name}'")
            workbook.create_sheet(sheet_name)
            return

        # Create a new worksheet
        logger.info(f"Creating worksheet '{sheet_name}' with {len(data)} rows")
        worksheet = workbook.create_sheet(sheet_name)

        # Write data to worksheet
        rows_written = 0
        for row_idx, row in enumerate(data, 1):
            try:
                for col_idx, cell_value in enumerate(row, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                rows_written += 1
            except Exception as e:
                logger.error(
                    f"Error writing row {row_idx} to worksheet '{sheet_name}': {e}"
                )

        # Adjust column widths based on content
        from openpyxl.utils import get_column_letter

        if data:  # Only adjust if we have data
            # Get the number of columns
            num_columns = len(data[0])

            # Add autofilter to the header row
            last_column_letter = get_column_letter(num_columns)
            worksheet.auto_filter.ref = f"A1:{last_column_letter}{len(data)}"
            logger.info(f"Added autofilter to header row in worksheet '{sheet_name}'")

            # Adjust column widths
            for col_idx in range(1, num_columns + 1):
                column_letter = get_column_letter(col_idx)
                # Get maximum content length in the column
                max_length = 0
                for row_idx in range(1, len(data) + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        # For multiline cells, get the longest line
                        if isinstance(cell_value, str) and "\n" in cell_value:
                            lines = cell_value.split("\n")
                            max_line_length = max(len(line) for line in lines)
                            max_length = max(max_length, max_line_length)
                        else:
                            max_length = max(max_length, len(str(cell_value)))

                # Add some padding and set the width (with a minimum width of 10)
                adjusted_width = max(10, max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(
            f"Successfully wrote {rows_written} rows to worksheet '{sheet_name}' with adjusted column widths and autofilter"
        )
    except Exception as e:
        logger.error(f"Error creating worksheet '{sheet_name}': {e}")
        logger.error(traceback.format_exc())


def initialize_report_data():
    """
    Initialize an empty data structure for report types.

    Returns:
        Dict: Empty data structure for document_export, document_import, and dossier_import
    """
    return {
        "document_export": [],
        "document_import": [],
        "dossier_import": [],
    }


def get_all_municipalities_in_segment(report_types: Dict) -> Set[str]:
    """
    Get all unique municipalities across all report types for a segment.

    Args:
        report_types: Dictionary with structure {report_type: {municipality: [csv_files]}}

    Returns:
        Set of unique municipality names
    """
    all_municipalities = set()
    for report_type in report_types.values():
        all_municipalities.update(report_type.keys())
    return all_municipalities


def create_empty_workbook() -> openpyxl.Workbook:
    """
    Create a new Excel workbook with the default worksheet removed.

    Returns:
        openpyxl.Workbook: A new workbook with no worksheets
    """
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    return workbook


def process_municipality_data(
    municipality: str,
    segment: str,
    report_types: Dict,
    segment_dir: Path,
    segment_data: Dict[str, List],
    all_data: Dict[str, List],
    oldest_import_dir: Optional[str],
) -> Tuple[int, int]:
    """
    Process data for a single municipality and create its Excel file.

    Args:
        municipality: Name of the municipality
        segment: Name of the segment
        report_types: Dictionary with structure {report_type: {municipality: [csv_files]}}
        segment_dir: Path to the segment directory
        segment_data: Dictionary to store data for the segment
        all_data: Dictionary to store data for the "all" workbook
        oldest_import_dir: Path to the oldest import directory (chronologically)

    Returns:
        Tuple of (total_excel_files, successful_files)
    """
    total_excel_files = 0
    successful_files = 0

    try:
        # Create individual Excel file for this municipality
        excel_path = segment_dir / f"{municipality}_{segment}.xlsx"
        logger.info(f"Creating Excel file: {excel_path}")
        total_excel_files += 1

        # Create Excel workbook for this municipality
        municipality_workbook = create_empty_workbook()

        # Process each report type
        for report_type in ["document_export", "document_import", "dossier_import"]:
            try:
                # Get CSV files for this municipality and report type
                csv_files = report_types.get(report_type, {}).get(municipality, [])

                # Process CSV files once
                data = process_csv_files(
                    csv_files, report_type, oldest_import_dir, segment
                )

                # Write to individual municipality Excel file
                write_to_excel(data, municipality_workbook, report_type)

                # Add to segment data (for merged segment workbook)
                if segment_data[report_type] and data and len(data) > 1:
                    # Skip header row for all but the first municipality
                    segment_data[report_type].extend(data[1:])
                else:
                    segment_data[report_type].extend(data)

                # Add to all data (for the "all" workbook) with Segment column
                if data:
                    if not all_data[report_type]:
                        # First data - add header with Segment column at position 2
                        if len(data) > 0:
                            header = data[0].copy()
                            # Insert "Segment" at position 1 (index 1, position 2)
                            header.insert(1, "Segment")
                            all_data[report_type].append(header)
                            # Add data rows with segment value
                            for row in data[1:]:
                                new_row = row.copy()
                                new_row.insert(1, segment)
                                all_data[report_type].append(new_row)
                    else:
                        # Subsequent data - skip header, add rows with segment
                        for row in data[1:]:
                            new_row = row.copy()
                            new_row.insert(1, segment)
                            all_data[report_type].append(new_row)

            except Exception as e:
                logger.error(
                    f"Error processing report type {report_type} for municipality {municipality}: {e}"
                )

        # Save the individual municipality workbook
        try:
            municipality_workbook.save(excel_path)
            successful_files += 1
            logger.info(f"Successfully saved Excel file: {excel_path}")
        except Exception as e:
            logger.error(f"Error saving Excel file {excel_path}: {e}")
    except Exception as e:
        logger.error(
            f"Error processing municipality {municipality} in segment {segment}: {e}"
        )

    return total_excel_files, successful_files


def create_segment_workbook(
    segment: str,
    segment_data: Dict[str, List],
    output_dir: Path,
) -> None:
    """
    Create and save a merged workbook for a segment.

    Args:
        segment: Name of the segment
        segment_data: Dictionary with data for each report type
        output_dir: Target directory for Excel files
    """
    # Create a workbook for this segment
    segment_workbook = create_empty_workbook()

    # Write segment data to the merged segment workbook
    for report_type, data in segment_data.items():
        write_to_excel(data, segment_workbook, report_type)

    # Save the merged segment workbook
    segment_excel_path = output_dir / f"{segment}.xlsx"
    try:
        segment_workbook.save(segment_excel_path)
        logger.info(
            f"Successfully saved merged Excel file for segment: {segment_excel_path}"
        )
    except Exception as e:
        logger.error(f"Error saving merged Excel file for segment {segment}: {e}")


def create_all_segments_workbook(
    all_data: Dict[str, List],
    output_dir: Path,
) -> None:
    """
    Create and save a workbook containing data from all segments.

    Args:
        all_data: Dictionary with data for each report type from all segments
        output_dir: Target directory for Excel files
    """
    try:
        logger.info("Creating 'all' workbook with all segments")

        # Create a workbook for all segments
        all_workbook = create_empty_workbook()

        # Write all data to workbook
        for report_type, data in all_data.items():
            write_to_excel(data, all_workbook, report_type)

        # Save the all workbook
        all_excel_path = output_dir / "all.xlsx"
        try:
            all_workbook.save(all_excel_path)
            logger.info(
                f"Successfully saved merged Excel file for all segments: {all_excel_path}"
            )
        except Exception as e:
            logger.error(f"Error saving merged Excel file for all segments: {e}")
    except Exception as e:
        logger.error(f"Error creating 'all' workbook: {e}")


def get_consolidated_rows_for_gemeinde(  # noqa: C901
    central_data: Dict,
    report_type: str,
    segment: str,
    gemeinde: str,
    headers: List[str],
) -> List[List[str]]:
    """
    Extract consolidated rows for a specific gemeinde in a segment.

    Args:
        central_data: The central data structure
        report_type: Type of report
        segment: Segment name
        gemeinde: Gemeinde name
        headers: Column headers

    Returns:
        List of consolidated rows including header
    """
    result = [headers]

    if report_type not in central_data:
        return result
    if segment not in central_data[report_type]:
        return result
    if gemeinde not in central_data[report_type][segment]:
        return result

    # Get error column index
    _, error_column_name = get_key_columns_for_report_type(report_type)
    try:
        error_column_idx = headers.index(error_column_name)
    except ValueError:
        logger.warning(
            f"Error column '{error_column_name}' not found for {report_type}"
        )
        error_column_idx = -1

    # Collect all keys and their consolidated rows for this gemeinde
    keys_and_rows = []

    for gesuch_id in sorted(central_data[report_type][segment][gemeinde].keys()):
        if report_type in ["document_export", "document_import"]:
            # Has dateiname level
            for dateiname in sorted(
                central_data[report_type][segment][gemeinde][gesuch_id].keys()
            ):
                rows = central_data[report_type][segment][gemeinde][gesuch_id][
                    dateiname
                ]
                key = (segment, gemeinde, gesuch_id, dateiname)
                consolidated_row = get_consolidated_row_from_list(
                    rows, error_column_idx
                )
                if consolidated_row:
                    keys_and_rows.append((key, consolidated_row))
        else:
            # dossier_import: no dateiname level
            rows = central_data[report_type][segment][gemeinde][gesuch_id]
            key = (segment, gemeinde, gesuch_id)
            consolidated_row = get_consolidated_row_from_list(rows, error_column_idx)
            if consolidated_row:
                keys_and_rows.append((key, consolidated_row))

    # Filter out keys with empty components if more specific keys exist
    keys_to_keep = set()
    for i, (key1, _) in enumerate(keys_and_rows):
        should_keep = True
        if has_empty_key_component(key1):
            # Check if there's a more specific key
            for j, (key2, _) in enumerate(keys_and_rows):
                if i != j and is_more_specific_key(key2, key1):
                    should_keep = False
                    break
        if should_keep:
            keys_to_keep.add(i)

    # Add kept rows to result
    for i in keys_to_keep:
        _, row = keys_and_rows[i]
        result.append(row)

    return result


def get_consolidated_rows_for_segment(  # noqa: C901
    central_data: Dict, report_type: str, segment: str, headers: List[str]
) -> List[List[str]]:
    """
    Extract consolidated rows for all gemeinden in a segment.

    Args:
        central_data: The central data structure
        report_type: Type of report
        segment: Segment name
        headers: Column headers

    Returns:
        List of consolidated rows including header
    """
    result = [headers]

    if report_type not in central_data:
        return result
    if segment not in central_data[report_type]:
        return result

    # Get error column index
    _, error_column_name = get_key_columns_for_report_type(report_type)
    try:
        error_column_idx = headers.index(error_column_name)
    except ValueError:
        logger.warning(
            f"Error column '{error_column_name}' not found for {report_type}"
        )
        error_column_idx = -1

    # Collect all keys and their consolidated rows for this segment
    keys_and_rows = []

    for gemeinde in sorted(central_data[report_type][segment].keys()):
        for gesuch_id in sorted(central_data[report_type][segment][gemeinde].keys()):
            if report_type in ["document_export", "document_import"]:
                # Has dateiname level
                for dateiname in sorted(
                    central_data[report_type][segment][gemeinde][gesuch_id].keys()
                ):
                    rows = central_data[report_type][segment][gemeinde][gesuch_id][
                        dateiname
                    ]
                    key = (segment, gemeinde, gesuch_id, dateiname)
                    consolidated_row = get_consolidated_row_from_list(
                        rows, error_column_idx
                    )
                    if consolidated_row:
                        keys_and_rows.append((key, consolidated_row))
            else:
                # dossier_import: no dateiname level
                rows = central_data[report_type][segment][gemeinde][gesuch_id]
                key = (segment, gemeinde, gesuch_id)
                consolidated_row = get_consolidated_row_from_list(
                    rows, error_column_idx
                )
                if consolidated_row:
                    keys_and_rows.append((key, consolidated_row))

    # Filter out keys with empty components if more specific keys exist
    keys_to_keep = set()
    for i, (key1, _) in enumerate(keys_and_rows):
        should_keep = True
        if has_empty_key_component(key1):
            # Check if there's a more specific key
            for j, (key2, _) in enumerate(keys_and_rows):
                if i != j and is_more_specific_key(key2, key1):
                    should_keep = False
                    break
        if should_keep:
            keys_to_keep.add(i)

    # Add kept rows to result
    for i in keys_to_keep:
        _, row = keys_and_rows[i]
        result.append(row)

    return result


def create_excel_workbooks_from_central_data(  # noqa: C901
    central_data: Dict,
    headers_by_report_type: Dict[str, List[str]],
    output_dir: Path,
):
    """
    Create all Excel files from the central data structure.

    - all.xlsx with all rows from all migrations
    - all-consolidated.xlsx with consolidated rows
    - segment-level Excel files (one per segment)
    - gemeinde-level Excel files (one per gemeinde per segment)

    Args:
        central_data: The central data structure
        headers_by_report_type: Dict mapping report_type to its headers
        output_dir: Target directory for Excel files
    """
    try:
        # Create output directory if it doesn't exist
        create_directory(output_dir)

        # Generate all.xlsx - contains all rows from all migrations
        logger.info("Generating all.xlsx with all rows from all migrations")
        all_workbook = create_empty_workbook()

        for report_type in ["document_export", "document_import", "dossier_import"]:
            if report_type not in headers_by_report_type:
                logger.warning(f"No headers found for {report_type}, skipping")
                continue

            headers = headers_by_report_type[report_type]
            logger.info(f"Processing {report_type} for all.xlsx")

            # Flatten to all rows
            all_rows = flatten_central_data_to_rows(central_data, report_type, headers)
            logger.info(f"  Flattened {len(all_rows) - 1} data rows for {report_type}")

            # Sort rows
            sorted_rows = sort_rows_by_keys(all_rows, report_type, headers)
            logger.info(f"  Sorted rows for {report_type}")

            # Write to Excel
            write_to_excel(sorted_rows, all_workbook, report_type)

        # Save all.xlsx
        all_xlsx_path = output_dir / "all.xlsx"
        try:
            all_workbook.save(all_xlsx_path)
            logger.info(f"Successfully saved {all_xlsx_path}")
        except Exception as e:
            logger.error(f"Error saving {all_xlsx_path}: {e}")

        # Generate all-consolidated.xlsx - contains consolidated rows
        logger.info("Generating all-consolidated.xlsx with consolidated rows")
        consolidated_workbook = create_empty_workbook()

        for report_type in ["document_export", "document_import", "dossier_import"]:
            if report_type not in headers_by_report_type:
                continue

            headers = headers_by_report_type[report_type]
            logger.info(f"Processing {report_type} for all-consolidated.xlsx")

            # Flatten to consolidated rows
            consolidated_rows = flatten_central_data_to_rows_consolidated(
                central_data, report_type, headers
            )
            logger.info(
                f"  Consolidated to {len(consolidated_rows) - 1} data rows for {report_type}"
            )

            # Sort rows
            sorted_rows = sort_rows_by_keys(consolidated_rows, report_type, headers)
            logger.info(f"  Sorted rows for {report_type}")

            # Write to Excel
            write_to_excel(sorted_rows, consolidated_workbook, report_type)

        # Save all-consolidated.xlsx
        consolidated_xlsx_path = output_dir / "all-consolidated.xlsx"
        try:
            consolidated_workbook.save(consolidated_xlsx_path)
            logger.info(f"Successfully saved {consolidated_xlsx_path}")
        except Exception as e:
            logger.error(f"Error saving {consolidated_xlsx_path}: {e}")

        # Generate segment-level and gemeinde-level Excel files
        logger.info("Generating segment-level and gemeinde-level Excel files")

        # Collect all unique segments across all report types
        all_segments = set()
        for report_type in ["document_export", "document_import", "dossier_import"]:
            if report_type in central_data:
                all_segments.update(central_data[report_type].keys())

        for segment in sorted(all_segments):
            # Create segment directory
            segment_dir = output_dir / segment
            create_directory(segment_dir)

            logger.info(f"Processing segment: {segment}")

            # Generate segment-level Excel file (consolidated)
            segment_workbook = create_empty_workbook()

            for rt in ["document_export", "document_import", "dossier_import"]:
                if rt not in headers_by_report_type:
                    continue

                headers = headers_by_report_type[rt]
                segment_rows = get_consolidated_rows_for_segment(
                    central_data, rt, segment, headers
                )

                if len(segment_rows) > 1:  # Has data beyond header
                    # Sort rows
                    sorted_segment_rows = sort_rows_by_keys(segment_rows, rt, headers)
                    write_to_excel(sorted_segment_rows, segment_workbook, rt)

            # Save segment-level Excel file at top level
            segment_excel_path = output_dir / f"{segment}.xlsx"
            try:
                segment_workbook.save(segment_excel_path)
                logger.info(f"  Saved segment Excel file: {segment_excel_path}")
            except Exception as e:
                logger.error(
                    f"  Error saving segment Excel file {segment_excel_path}: {e}"
                )

            # Collect all unique gemeinden in this segment across all report types
            all_gemeinden = set()
            for report_type in ["document_export", "document_import", "dossier_import"]:
                if report_type in central_data and segment in central_data[report_type]:
                    all_gemeinden.update(central_data[report_type][segment].keys())

            # Generate gemeinde-level Excel files (consolidated)
            for gemeinde in sorted(all_gemeinden):
                logger.info(f"  Processing gemeinde: {gemeinde}")

                gemeinde_workbook = create_empty_workbook()

                for rt in ["document_export", "document_import", "dossier_import"]:
                    if rt not in headers_by_report_type:
                        continue

                    headers = headers_by_report_type[rt]
                    gemeinde_rows = get_consolidated_rows_for_gemeinde(
                        central_data, rt, segment, gemeinde, headers
                    )

                    if len(gemeinde_rows) > 1:  # Has data beyond header
                        # Sort rows
                        sorted_gemeinde_rows = sort_rows_by_keys(
                            gemeinde_rows, rt, headers
                        )
                        write_to_excel(sorted_gemeinde_rows, gemeinde_workbook, rt)

                # Save gemeinde-level Excel file
                gemeinde_excel_path = segment_dir / f"{gemeinde}_{segment}.xlsx"
                try:
                    gemeinde_workbook.save(gemeinde_excel_path)
                    logger.info(f"    Saved gemeinde Excel file: {gemeinde_excel_path}")
                except Exception as e:
                    logger.error(
                        f"    Error saving gemeinde Excel file {gemeinde_excel_path}: {e}"
                    )

        logger.info("Excel file creation complete.")
    except Exception as e:
        logger.error(f"Error in create_excel_workbooks_from_central_data: {e}")
        logger.error(traceback.format_exc())


def main():
    """
    Orchestrate the consolidation process.

    The function performs the following steps:
    1. Parse command line arguments
    2. Parse all CSV files chronologically into centralized data structure
    3. Generate all.xlsx with all rows from all migrations (sorted)
    4. Generate all-consolidated.xlsx with consolidated rows (first error-free or last, with specificity filtering, sorted)
    5. Generate segment-level Excel files (one per segment, consolidated)
    6. Generate gemeinde-level Excel files (one per gemeinde per segment, consolidated)

    Returns:
        int: Exit code (0 for success, 1 for error)
    """
    try:
        logger.info("Starting consolidation of migration reports")
        start_time = __import__("time").time()

        # Parse command line arguments
        try:
            args = parse_arguments()
            logger.info(f"Input paths: {args.input_paths}")
            logger.info(f"Output directory: {args.output}")
        except Exception as e:
            logger.error(f"Error parsing command line arguments: {e}")
            return 1

        # Parse all CSV files into central data structure
        try:
            logger.info("Parsing all CSV files into central data structure")
            central_data, headers_by_report_type = (
                parse_all_csv_files_to_central_structure(args.input_paths)
            )

            # Check if any data was found
            if not any(central_data.values()):
                logger.warning("No data was found in the specified input paths")
                return 0

            logger.info("Successfully parsed all CSV files into central data structure")
            for report_type in ["document_export", "document_import", "dossier_import"]:
                if report_type in central_data and central_data[report_type]:
                    segment_count = len(central_data[report_type])
                    logger.info(f"  {report_type}: {segment_count} segments")
        except Exception as e:
            logger.error(f"Error parsing CSV files: {e}")
            logger.error(traceback.format_exc())
            return 1

        # Create Excel files
        try:
            logger.info("Creating Excel files from central data structure")
            # Convert output path to absolute path if it's relative
            output_path = Path(args.output).resolve()

            # Create all.xlsx and all-consolidated.xlsx
            create_excel_workbooks_from_central_data(
                central_data, headers_by_report_type, output_path
            )
        except Exception as e:
            logger.error(f"Error creating Excel files: {e}")
            logger.error(traceback.format_exc())
            return 1

        # Calculate execution time
        execution_time = __import__("time").time() - start_time
        logger.info(
            f"Consolidation complete! Execution time: {execution_time:.2f} seconds"
        )
        return 0
    except Exception as e:
        logger.error(f"Unexpected error in main function: {e}")
        logger.error(traceback.format_exc())
        return 1


if __name__ == "__main__":
    # Example regex filters (can be modified as needed)
    # DOCUMENT_EXPORT_FILTERS = [re.compile(r"pattern1"), re.compile(r"pattern2")]
    # DOCUMENT_IMPORT_FILTERS = [re.compile(r"pattern3")]
    # DOSSIER_IMPORT_FILTERS = [re.compile(r"pattern4")]

    # Set up example filters - uncomment and modify as needed
    # DOCUMENT_EXPORT_FILTERS = [
    #     re.compile(r"error", re.IGNORECASE),
    #     re.compile(r"failed", re.IGNORECASE)
    # ]
    # DOCUMENT_IMPORT_FILTERS = [
    #     re.compile(r"rejected", re.IGNORECASE)
    # ]
    # DOSSIER_IMPORT_FILTERS = [
    #     re.compile(r"invalid", re.IGNORECASE)
    # ]

    sys.exit(main())

# Summary of script functionality:
# 1. Accepts one or more input paths to migration report directories
# 2. Sorts input directories chronologically (oldest first)
# 3. Creates a target directory structure for consolidated Excel files
# 4. For each segment found in the input directories:
#    - Creates a segment directory in the output
#    - For each municipality found in the segment:
#      - Creates an Excel file with three worksheets (document_export, document_import, dossier_import)
#      - Combines data from all CSV files for that municipality and segment
#      - Applies filters based on chronological order:
#        * For the oldest import directory: applies both first-time and always filters for each report type
#        * For other directories: applies only always filters for each report type
#        * If no always filters are defined for a report type, no filtering is applied to non-oldest directories
# 5. Creates merged Excel workbooks at the root level:
#    - One workbook per segment containing all municipalities for that segment
#    - One "all" workbook containing all segments
# 6. Preserves multi-line cells in the Excel output
# 7. Provides detailed logging of the consolidation process
