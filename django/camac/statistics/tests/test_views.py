import datetime
import io
import os
from decimal import Decimal
from unittest.mock import MagicMock

import openpyxl
import pyexcel
import pytest
from caluma.caluma_form.models import Form
from django.urls import reverse
from pytest_lazy_fixtures import lf
from rest_framework import status

from camac.conftest import parse_xlsx_response
from camac.statistics.filters import (
    InstanceFilterBackend,
    WorkItemFilterBackend,
)
from camac.statistics.views import (
    DossierStatisticsExportView,
    _resolve_template_path,
)
from camac.work_items.models import WorkItemTemplate

STATISTICS_URL = reverse("statistics-dossiers")
WORK_ITEMS_URL = reverse("statistics-work-items")
EXPECTED_WORK_ITEM_COLUMNS = 15


@pytest.mark.parametrize(
    "role__name,setup_instance,expected_status,expected_rows",
    [
        ("Municipality", lf("statistics_ag_instance"), status.HTTP_200_OK, 2),
        (
            "Municipality",
            lf("statistics_ag_instance_by_role"),
            status.HTTP_200_OK,
            2,
        ),
        ("Service", lf("statistics_ag_instance_afb"), status.HTTP_200_OK, 2),
        ("Applicant", lf("statistics_ag_instance"), status.HTTP_403_FORBIDDEN, None),
    ],
)
def test_statistics_export_ag(
    admin_client,
    setup_instance,
    expected_status,
    expected_rows,
    snapshot,
):
    response = admin_client.get(STATISTICS_URL)

    assert response.status_code == expected_status

    if expected_status != status.HTTP_200_OK:
        return

    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    book = parse_xlsx_response(response)
    sheet = book.get_dict()["Data"]

    assert len(sheet) == expected_rows

    if expected_rows > 1:
        assert sheet[1] == snapshot


@pytest.mark.parametrize("role__name", ["Municipality"])
def test_statistics_export_empty(
    admin_client,
    db,
    ag_distribution_settings,
    multilang,
    settings,
):
    """Without any instance data the export should return only the header row."""
    settings.APPLICATION_NAME = "kt_ag"

    response = admin_client.get(STATISTICS_URL)

    assert response.status_code == status.HTTP_200_OK
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    book = parse_xlsx_response(response)
    sheet = book.get_dict()["Data"]
    assert len(sheet) == 1
    assert len(sheet[0]) >= 6  # at least default dossier columns


@pytest.mark.parametrize("role__name", ["Service"])
def test_work_items_export(
    admin_client,
    statistics_ag_instance_afb,
    snapshot,
):
    """Work-items export returns completed inquiry work items with all columns."""
    response = admin_client.get(WORK_ITEMS_URL)

    assert response.status_code == status.HTTP_200_OK
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="work-items_service-afb.xlsx"' in response["Content-Disposition"]

    book = parse_xlsx_response(response)
    sheet = book.get_dict()["Data"]

    # Header + 1 completed inquiry (decision + other tasks are excluded)
    assert len(sheet) == 2
    assert len(sheet[0]) == EXPECTED_WORK_ITEM_COLUMNS

    assert sheet == snapshot

    # Date columns must be native datetime cells, not strings.
    date_labels = {
        "Eingabedatum",
        "Eingegangen",
        "Abgeschlossen",
        "Erstellt am",
        "Frist",
        "Abgeschlossen am",
    }
    header_row, data_row = sheet[0], sheet[1]
    for idx, label in enumerate(header_row):
        if label in date_labels and data_row[idx]:
            assert isinstance(data_row[idx], (datetime.date, datetime.datetime)), (
                f"'{label}' should be a date, got {type(data_row[idx])}"
            )


@pytest.mark.parametrize("role__name", ["Service"])
def test_statistics_service_afb_deadline_columns(
    admin_client,
    statistics_ag_instance_afb,
    snapshot,
):
    """service-afb dossier export includes processing_time and on_time columns."""
    response = admin_client.get(STATISTICS_URL)

    assert response.status_code == status.HTTP_200_OK

    book = parse_xlsx_response(response)
    sheet = book.get_dict()["Data"]

    assert len(sheet) == 2  # header + 1 data row
    assert len(sheet[0]) == 10  # 8 default + processing_time + on_time

    assert sheet == snapshot


@pytest.mark.parametrize("role__name", ["Municipality"])
def test_copy_template_sheets(
    admin_client,
    statistics_ag_instance,
    tmp_path,
    settings,
    mocker,
    freezer,
    snapshot,
):
    """Verify that it copies extra sheets from a template and adds metadata."""
    freezer.move_to("2026-08-12")

    # minimal template with a "Data" sheet and a "Pivot" sheet.
    tpl_path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    ws_pivot = wb.create_sheet("Pivot")
    ws_pivot.cell(row=1, column=1, value="pivot-header")
    ws_pivot.cell(row=2, column=1, value=42)
    wb.save(tpl_path)

    mocker.patch(
        "camac.statistics.views._resolve_template_path",
        return_value=str(tpl_path),
    )

    response = admin_client.get(STATISTICS_URL, {"form": "baugesuch"})
    assert response.status_code == status.HTTP_200_OK

    content = b"".join(response.streaming_content)
    book = pyexcel.get_book(file_content=content, file_type="xlsx")
    assert sorted(book.to_dict().keys()) == snapshot

    exported_string = "Exportiert am: 12.08.2026"

    result = openpyxl.load_workbook(io.BytesIO(content))
    assert result["Pivot"].cell(row=2, column=1).value == exported_string
    assert result["Filter"].cell(row=2, column=1).value != exported_string
    assert result["Data"].cell(row=2, column=1).value != exported_string


@pytest.mark.parametrize(
    "export_type,expected_base",
    [
        ("dossiers", "dossiers"),
        ("work_items", "work-items"),
        ("billings", "billings"),
        ("rpg2", "rpg2"),
        ("unknown_custom", "unknown-custom"),  # fallback: no base type matches
    ],
)
@pytest.mark.parametrize(
    "available,expected_suffix",
    [
        ({"sg", "role", "default"}, "sg"),
        ({"role", "default"}, "role"),
        ({"default"}, "default"),
        (set(), None),
    ],
)
def test_resolve_template_priority(
    settings,
    mocker,
    export_type,
    expected_base,
    available,
    expected_suffix,
):
    """Template resolution prefers service-group > role > type-default."""

    settings.APPLICATION_NAME = "kt_ag"
    base = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "config",
        "kt_ag",
    )

    paths = {
        "sg": f"{base}/{expected_base}_service-afb.xlsx",
        "role": f"{base}/{expected_base}_municipality.xlsx",
        "default": f"{base}/{expected_base}_export.xlsx",
    }

    existing = {paths[k] for k in available}
    mocker.patch("os.path.isfile", side_effect=lambda p: p in existing)

    expected = paths[expected_suffix] if expected_suffix else None
    assert (
        _resolve_template_path(export_type, "municipality", "service-afb") == expected
    )


def test_write_data_sheet_non_native_types_and_excess_columns():
    """Non-native types are str-converted; excess columns in rows are ignored."""
    view = DossierStatisticsExportView()
    wb = openpyxl.Workbook()

    header = ["Col1", "Col2"]
    rows = [
        # Decimal is not a native type -> str()
        #  third value exceeds header -> skipped
        (Decimal("3.14"), "normal", "excess"),
    ]

    view._write_data_sheet(wb, header, rows)

    sheet = wb["Data"]
    assert sheet.cell(row=2, column=1).value == "3.14"
    assert sheet.cell(row=2, column=2).value == "normal"
    assert sheet.cell(row=2, column=3).value is None  # not written


def test_set_pivot_refresh():
    """Pivot caches on the workbook get refreshOnLoad=True."""
    cache = MagicMock()
    cache.refreshOnLoad = False

    wb = MagicMock()
    wb._pivot_caches = [cache]

    DossierStatisticsExportView._set_pivot_refresh(wb)
    assert cache.refreshOnLoad is True


def test_build_workbook_without_template():
    """When no template is given, _build_workbook creates a fresh workbook."""
    view = DossierStatisticsExportView()
    buf = view._build_workbook(["A", "B"], [("x", "y")], template_path=None)
    wb = openpyxl.load_workbook(buf)
    assert wb["Data"].cell(row=1, column=1).value == "A"
    assert wb["Data"].cell(row=2, column=1).value == "x"


def test_write_filter_sheet_replaces_existing():
    """_write_filter_sheet removes a pre-existing 'Filter' sheet."""
    wb = openpyxl.Workbook()
    wb.create_sheet("Filter")
    assert "Filter" in wb.sheetnames

    view = DossierStatisticsExportView()
    view._write_filter_sheet(wb, [("Param", "Value")])

    assert wb.sheetnames[0] == "Filter"
    assert wb["Filter"].cell(row=4, column=1).value == "Param"


def test_dossier_annotations_all_columns(db, settings, snapshot):
    """Calling _dossier_annotations without requested_columns returns all."""
    settings.APPLICATION_NAME = "kt_ag"

    backend = InstanceFilterBackend()
    annotations = backend._dossier_annotations(service_id=1, requested_columns=None)

    assert sorted(annotations.keys()) == snapshot


def test_work_item_annotations_all_columns(db, settings, snapshot):
    """Calling _work_item_annotations without requested_columns returns all."""
    settings.APPLICATION_NAME = "kt_ag"

    backend = WorkItemFilterBackend()
    annotations = backend._work_item_annotations(requested_columns=None)

    assert sorted(annotations.keys()) == snapshot


@pytest.mark.parametrize("role__name", ["Municipality"])
def test_filter_sheet(
    admin_client,
    statistics_ag_instance,
    caluma_option_factory,
    freezer,
):
    """Export with query filters produces a Filter sheet with applied filters."""
    freezer.move_to("2026-04-10")

    Form.objects.filter(slug="baugesuch").update(name={"de": "Baugesuch"})
    caluma_option_factory(slug="test-approved", label="Bewilligt")
    inst_state_id = str(statistics_ag_instance.instance_state_id)

    response = admin_client.get(
        STATISTICS_URL,
        {
            "submit_date_after": "2025-01-01",
            "form": "baugesuch",
            "instance_state": inst_state_id,
            "decision": "test-approved",
        },
    )
    assert response.status_code == status.HTTP_200_OK

    wb = openpyxl.load_workbook(io.BytesIO(b"".join(response.streaming_content)))
    assert wb.sheetnames[0] == "Filter"
    assert "Data" in wb.sheetnames

    fs = wb["Filter"]
    assert fs.cell(row=1, column=1).value == "Übersicht der Filter"

    # Rows 2-3 are blank, filters start at row 4.
    assert fs.cell(row=4, column=1).value == "Eingabedatum von"
    assert fs.cell(row=4, column=2).value == "01.01.2025"
    assert fs.cell(row=5, column=1).value == "Gesuchstyp"
    assert fs.cell(row=5, column=2).value == "Baugesuch"
    assert fs.cell(row=6, column=1).value == "Dossier-Status"
    assert fs.cell(row=6, column=2).value == "In Zirkulation"
    assert fs.cell(row=7, column=1).value == "Bauentscheid"
    assert fs.cell(row=7, column=2).value == "Bewilligt"

    # One blank row, then "Exportiert am".
    assert fs.cell(row=9, column=1).value == "Exportiert am"
    assert fs.cell(row=9, column=2).value == datetime.datetime(2026, 4, 10, 0, 0)


@pytest.mark.parametrize("role__name", ["Service"])
def test_filter_sheet_work_items(
    admin_client,
    statistics_ag_instance_afb,
    caluma_work_item_factory,
    group,
    freezer,
):
    """Work-items export with filters shows them in the Filter sheet."""
    freezer.move_to("2026-04-10")

    template = WorkItemTemplate.objects.create(
        name="Stellungnahme / Entscheid schreiben",
        responsibility_rule="NONE",
    )
    caluma_work_item_factory(
        case=statistics_ag_instance_afb.case,
        task_id="create-manual-workitems",
        status="completed",
        addressed_groups=[str(group.service.pk)],
        meta={"template-id": str(template.pk)},
    )

    response = admin_client.get(
        WORK_ITEMS_URL,
        {
            "wi_created_at_after": "2025-01-01",
            "task": f"inquiry,{template.pk},unknown-slug",
        },
    )
    assert response.status_code == status.HTTP_200_OK

    content = b"".join(response.streaming_content)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames[0] == "Filter"

    fs = wb["Filter"]
    assert fs.cell(row=4, column=1).value == "Aufgabe"
    assert fs.cell(row=4, column=2).value == (
        "Stellungnahme zustellen, Stellungnahme / Entscheid schreiben, unknown-slug"
    )
    assert fs.cell(row=5, column=1).value == "Aufgabe erstellt von"
    assert fs.cell(row=5, column=2).value == "01.01.2025"
