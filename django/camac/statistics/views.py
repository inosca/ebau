import io
import os
import uuid
from datetime import date, datetime, time, timedelta

import openpyxl
from caluma.caluma_form.models import Form, Option
from caluma.caluma_workflow.models import Task
from django.conf import settings
from django.http import FileResponse
from django.utils import timezone
from django.utils.translation import get_language, gettext_lazy as _
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import PermissionDenied
from rest_framework.generics import ListAPIView

from camac.instance.mixins import InstanceQuerysetMixin
from camac.instance.models import Instance, InstanceStateT
from camac.work_items.models import WorkItemTemplate

from .filters import InstanceFilterBackend, WorkItemFilterBackend

# Translatable strings used in the filter sheet.
_FILTER_OVERVIEW_TITLE = _("Filter overview")
_EXPORTED_ON_LABEL = _("Exported on")

_DATE_FORMAT = "DD.MM.YYYY"
_DATE_STRFTIME = "%d.%m.%Y"

# Maps query parameter names to translatable labels for the filter sheet.
FILTER_LABELS = [
    ("submit_date_after", _("Submission date from")),
    ("submit_date_before", _("Submission date to")),
    ("form", _("Application Type")),
    ("instance_state", _("Instance state")),
    ("decision", _("Decision")),
    ("first_inquiry_date_after", _("First inquiry date from")),
    ("first_inquiry_date_before", _("First inquiry date to")),
    ("completing_date_after", _("Completing date from")),
    ("completing_date_before", _("Completing date to")),
    ("involved", _("Involved")),
    ("task", _("Task")),
    ("wi_created_at_after", _("Work item created from")),
    ("wi_created_at_before", _("Work item created to")),
    ("wi_closed_at_after", _("Work item closed from")),
    ("wi_closed_at_before", _("Work item closed to")),
]


def _resolve_template_path(export_type, role_slug=None, service_group_slug=None):
    """Return the best-matching XLSX template path, or None.

    Priority: service_group > role > type-only default.
    """

    # normalize export_type _ to - to get base of template name
    base_type = export_type.replace("_", "-")

    stats_dir = os.path.join(
        os.path.dirname(__file__),
        "config",
        settings.APPLICATION_NAME,
    )

    candidates = []
    if service_group_slug:
        candidates.append(f"{base_type}_{service_group_slug}.xlsx")
    if role_slug:
        candidates.append(f"{base_type}_{role_slug}.xlsx")
    candidates.append(f"{base_type}_export.xlsx")

    for filename in candidates:
        path = os.path.join(stats_dir, filename)
        if os.path.isfile(path):
            return path

    return None


class _StatisticsExportBaseView(InstanceQuerysetMixin, ListAPIView):
    """Base class for statistics XLSX exports.

    Subclasses set base_type to control which
    columns / filter backend / template are used.

    The first worksheet of the template is filled with the queried data.
    Further worksheets (e.g. pivot tables) are preserved as-is from the
    template file.
    """

    # Overridden by concrete subclasses.
    base_type: str  # e.g. "dossier" or "work-items"

    instance_field = None
    queryset = Instance.objects
    filter_backends = [InstanceFilterBackend]

    def get_queryset_for_public(self):  # pragma: no cover
        return self.queryset.none()

    def _get_columns(self, request):
        """Resolve export columns from the statistics module config.

        Resolution order (first match wins): by_service_group, by_role.
        Raises PermissionDenied if no configuration matches.
        """
        stats_config = settings.STATISTICS
        columns_config = None

        service = getattr(request.group, "service", None)
        sg_slug = getattr(service.service_group, "slug", None) if service else None
        sg_columns = stats_config.by_service_group.get(sg_slug) if sg_slug else None
        if sg_columns:
            columns_config = sg_columns
        else:
            role = getattr(request.group, "role", None)
            role_slug = getattr(role, "slug", None) if role else None
            role_columns = stats_config.by_role.get(role_slug) if role_slug else None
            if role_columns:
                columns_config = role_columns

        if columns_config is None:
            raise PermissionDenied()

        return getattr(columns_config, self.base_type)

    def _get_template(self, request):
        """Resolve the XLSX template for the current export."""
        service = getattr(request.group, "service", None)
        sg_slug = getattr(service.service_group, "slug", None) if service else None
        role = getattr(request.group, "role", None)
        role_slug = getattr(role, "slug", None) if role else None

        return _resolve_template_path(self.base_type, role_slug, sg_slug)

    def _write_data_sheet(self, workbook, header, rows):
        """Clear and rewrite the "Data" sheet with *header* and *rows*."""
        if "Data" in workbook.sheetnames:
            del workbook["Data"]
        data_sheet = workbook.create_sheet("Data", 0)

        bold_font = Font(bold=True)
        col_widths = [0] * len(header)

        for col_idx, value in enumerate(header, start=1):
            cell = data_sheet.cell(row=1, column=col_idx, value=value)
            cell.font = bold_font
            col_widths[col_idx - 1] = len(str(value))

        _NATIVE_TYPES = (str, int, float, bool, datetime, date, time, timedelta)
        num_cols = len(header)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row):
                if col_idx >= num_cols:
                    break
                if value is not None and not isinstance(value, _NATIVE_TYPES):
                    value = str(value)
                cell = data_sheet.cell(row=row_idx, column=col_idx + 1, value=value)
                if isinstance(value, (date, datetime)):
                    cell.number_format = _DATE_FORMAT
                col_widths[col_idx] = max(col_widths[col_idx], len(str(value or "")))

        for col_idx, width in enumerate(col_widths):
            col_letter = get_column_letter(col_idx + 1)
            data_sheet.column_dimensions[col_letter].width = min(width + 2, 60)

    @staticmethod
    def _set_pivot_refresh(workbook):
        """Set refreshOnLoad on pivot caches (openpyxl private API)."""
        for cache in getattr(workbook, "_pivot_caches", []):
            cache.refreshOnLoad = True

    @staticmethod
    def _resolve_filter_value(param, raw_value):
        """Resolve raw filter values to human-readable labels where possible."""
        lang = get_language()
        values = [v.strip() for v in raw_value.split(",") if v.strip()]

        if param.endswith(("_after", "_before")):
            try:
                return date.fromisoformat(raw_value).strftime(_DATE_STRFTIME)
            except ValueError:  # pragma: no cover
                return raw_value

        if param == "form":
            forms = {
                form.slug: str(form.name)
                for form in Form.objects.filter(slug__in=values)
            }
            return ", ".join(forms.get(value, value) for value in values)

        if param == "instance_state":
            names = list(
                InstanceStateT.objects.filter(
                    instance_state_id__in=values,
                    language=lang,
                ).values_list("name", flat=True)
            )
            return ", ".join(names) if names else raw_value

        if param == "decision":
            options = Option.objects.filter(slug__in=values)
            labels = [str(opt.label) for opt in options]
            return ", ".join(labels) if labels else raw_value

        if param == "task":
            # values can be either Caluma task slugs or WorkItemTemplate UUIDs
            # (same dual lookup as in WorkItemFilterBackend.filter_queryset)
            uuid_values = []
            slug_values = []
            for value in values:
                try:
                    uuid_values.append(uuid.UUID(value))
                except ValueError:
                    slug_values.append(value)

            tasks = {
                task.slug: str(task.name)
                for task in Task.objects.filter(slug__in=slug_values)
            }
            templates = {
                str(tpl.pk): tpl.name
                for tpl in WorkItemTemplate.objects.filter(pk__in=uuid_values)
            }

            labels = [
                tasks.get(value) or templates.get(value) or value for value in values
            ]
            return ", ".join(labels)

        return raw_value  # pragma: no cover

    @staticmethod
    def _collect_applied_filters(request):
        """Return a list of (label, value) for every set query parameter."""
        filters = []
        for param, label in FILTER_LABELS:
            value = request.query_params.get(param, "").strip()
            if value:
                resolved = _StatisticsExportBaseView._resolve_filter_value(param, value)
                filters.append((str(label), resolved))
        return filters

    def _write_filter_sheet(self, workbook, filters):
        """Create a "Filter" sheet at position 0 listing applied filters."""
        if "Filter" in workbook.sheetnames:
            del workbook["Filter"]
        sheet = workbook.create_sheet("Filter", 0)

        bold = Font(bold=True)
        sheet.cell(row=1, column=1, value=str(_FILTER_OVERVIEW_TITLE)).font = bold

        row = 4  # two blank rows after the title
        for label, value in filters:
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=value)
            row += 1

        row += 1  # one blank row before "Exported on"
        sheet.cell(row=row, column=1, value=str(_EXPORTED_ON_LABEL))
        today = timezone.now().date()
        cell = sheet.cell(row=row, column=2, value=today)
        cell.number_format = _DATE_FORMAT

        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 30

    @staticmethod
    def _write_template_sheet_metadata(workbook, filters):
        """Write export metadata into row 2 and 3 of every template sheet."""
        today = timezone.now().date()
        created_text = f"{_EXPORTED_ON_LABEL}: {today.strftime('%d.%m.%Y')}"
        filter_text = "; ".join(f"{label}: {value}" for label, value in filters)

        for sheet_name in workbook.sheetnames:
            if sheet_name in ("Filter", "Data"):
                continue
            sheet = workbook[sheet_name]
            sheet.cell(row=2, column=1, value=created_text)
            sheet.cell(row=3, column=1, value=filter_text)

    def _build_workbook(self, header, rows, template_path, filters=None):
        """Return a BytesIO buffer with the finished XLSX workbook."""
        if template_path and os.path.isfile(template_path):
            workbook = openpyxl.load_workbook(template_path)
        else:
            workbook = openpyxl.Workbook()

        self._write_data_sheet(workbook, header, rows)
        self._write_filter_sheet(workbook, filters or [])
        self._write_template_sheet_metadata(workbook, filters or [])
        self._set_pivot_refresh(workbook)

        buf = io.BytesIO()
        workbook.save(buf)
        buf.seek(0)
        return buf

    def get(self, request):
        columns = self._get_columns(request)
        backend = self._get_filter_backend()

        annotation_names = [col for col, _ in columns]
        header = [str(label) for _, label in columns]

        queryset = backend.filter_queryset(
            request, self.get_queryset(), set(annotation_names)
        )

        rows = list(queryset.values_list(*annotation_names))

        filters = self._collect_applied_filters(request)
        template_path = self._get_template(request)
        buf = self._build_workbook(header, rows, template_path, filters)

        return FileResponse(
            buf,
            as_attachment=True,
            filename=os.path.basename(template_path)
            if template_path
            else f"{self.base_type.replace('_', '-')}_export.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class DossierStatisticsExportView(_StatisticsExportBaseView):
    """Export one row per matching instance (dossier)."""

    base_type = "dossiers"

    def _get_filter_backend(self):
        return InstanceFilterBackend()


class WorkItemStatisticsExportView(_StatisticsExportBaseView):
    """Export one row per completed work item."""

    base_type = "work_items"

    def _get_filter_backend(self):
        return WorkItemFilterBackend()
